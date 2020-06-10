# print("\n",file=sys.stdout)
# sourceFile = open('python.txt', 'w')
# print("Круто же, правда?", file = sourceFile)
# sourceFile.close()

# ================
# Flags (switches)
# ================
table_format = True
file_output = True
Excel_output = False
BandsFilter = True
fgi_out = True
geran_out = True
utran_out = True
utrangeranbinary = False
R14_enabled = True

# VARs
TMF = False
eNB = False
S256DL = False # 256QAM DL for some of use band(s)
S256UL = False # 256QAM UL for some of use band(s)
S_4x4 = False # 4 layers for some of use band(s)
S_8x8 = False # 8 layers for some of use band(s)
S_CA7C = False # CA 7C support (CA of 2 carriers in Band 7)
S_Qualcomm = False # Qualcomm's requirement statement value

# Bands filters and XLS conclusions base
PrimaryEUTRABand = 7 # Carrier of LTE band for which "Supported" decision should be made about 256QAM or 4/8 layers supporting
SecondaryEUTRABands = [3,7,20] # Carriers of LTE Bands, which combinations are supported in MF and lab ERANs for special marked output

# internal DB arrays:
Bands = []
CCs = []
CCsCC = []

# Temporary variables
temp_int = 0
temp_str = ""

# Output lists
UEaccS = []
UEcats = []
eUtraBands = []
UtraBands = []
GeRANBands = []
Combinations = []
GeranCScapTxt = ""
GeranPScapTxt = ""
UTRANcapTxt = ""
FGI8txt = ""
FGI9txt = ""
FGI10txt = ""

from  Patterns import *
import os
import openpyxl
import logging
from openpyxl.styles import Font, Fill  # Подключаем стили для текста
from openpyxl.styles import colors  # Подключаем цвета для текста и ячеек
from openpyxl.styles import PatternFill  # Подключаем стили для ячеек

# Настройка  модуля Logging
formatter = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
logging.basicConfig(level=logging.INFO, format=formatter)
logger = logging.getLogger('_Parser')
logger.setLevel(logging.INFO)

# извлечение числа из строчного значения в скобках вида (25)
def valfrombrackets(inputstr):
    """
    This for parse something
    :param inputstr: this is regular expression
    :return: string value withour brackets
    e.g. inputstr = '(bla vla)'
    return = 'bla vla'
    """
    pos01 = inputstr.find("(")
    pos02 = inputstr.find(")")
    if (pos01 >= 0 and pos02 > 0):
        return inputstr[pos01 + 1:pos02]
    else:
        logger.error('ValFromBrackets cannot find any value in %s:', inputstr)
        logger.error('Returned None')
        return None

def valfromGSM(inputstr):
    """
    This for parse something
    :param inputstr: this is regular expression
    :return: INTEGER value of gsm frequency
    e.g. inputstr = 'SupportedBandGERAN :  ---- gsm900E(7) ---- 00111***'
    return = 900
    """
    aa = ''
    pos01 = inputstr.find("gsm")
    pos02 = inputstr.find("(")
    if (pos01 >= 0 and pos02 > 0):
        a = inputstr[pos01+3:pos02]
        for b in range(0, len(a)):
            if (a[b] in '0123456789'): aa= aa+a[b]
        if len(a)>0:
            return int(aa)
        else:
            return None
    else:
        logger.error('ValFromGSM cannot find any value in %s:', inputstr)
        logger.error('Returned None')
        return None

def EndStrFilter(inputstr):
    """
    This for parse something
    :param inputstr: this is regular expression
    :return: inputstr up to ')' symbol without /n
    e.g. inputstr = 'ue-Category --- 0x4(4)']'
    return = 'ue-Category --- 0x4(4)'
    """
    pos01 = inputstr.find(")")
    if (pos01 > 0):
        a = inputstr[:pos01+1]
        a = a.rstrip('/n')
        if len(a)>0:
            return a
        else:
            return ''
    else:
        logger.error('EndStrFilter() cannot find any value in %s:', inputstr)
        logger.error('Returned None')
        return ''

# Call dialogue box to ask input file with capability_info trace
# import tkinter as tk
import tkinter.filedialog
import sys
from prettytable import PrettyTable

t1 = PrettyTable()
t2 = PrettyTable()

if len(sys.argv) < 2:  # нет параметров
    fn = tkinter.filedialog.askopenfilename()
else:
    fn = sys.argv[1]
cwd = os.getcwd()  # текущий каталог
if os.path.exists(fn):  # если файл трассировки существует
    f = open(fn, 'r')
    Fname, ext = os.path.splitext(fn)
    # считываем текстовый файл в список
    # Should be '.txt'
    if ('txt' in ext):
        # Add prefix 'Parsed_' to the input file name for the ouput (parsed) filename (with the same file path)
        fnO = os.path.join(os.path.split(fn)[0], 'Parsed_' + os.path.split(fn)[1])
        logging.info('Output text file: %s', fnO)
        s = f.readlines()
        f.close()
    # Should be '.xls','.xlsx','.xlsm'
    elif ('xls' in ext):
        Excel_output = True
        XLwithmacro = ('.xlsm','.xltm')
        XLwithoutmacro = ('.xls','.xlsx','.xltx')
        # меняем логгер на XLS
        logger = logging.getLogger('XLSproc')
        # Входная трассировка в файле Excel на вкладке Capabilities
        logger.info('Open workbook: %s', fn)
        if (ext in XLwithmacro):
            book = openpyxl.load_workbook(fn, keep_vba=True)
        elif (ext in XLwithoutmacro):
            book = openpyxl.load_workbook(fn)
        else:
            logging.error('Extension %s is not known', ext)
            exit(1)
        # Add prefix 'Parsed_' to the input file name for the ouput (parsed) filename (with the same file path)
        # Output Text file name
        # FnameX, extX = os.path.splitext(fn0)
        # Output XLS file name with the same extension as input file
        # fnX = os.path.join(os.path.split(fn)[0], 'Parsed_' + os.path.split(fn)[1])
        fnX = os.path.splitext(fn)[0]+'_parsed'+ext
        logging.info('Output XLS file: %s', fnX)
        # Output TXT file name is the same like input filename, but with '_parsed' suffix and '.txt' extension
        fnO = os.path.splitext(fn)[0]+'_parsed.txt'
        logging.info('Output text file: %s', fnO)
        logger.info('Loading device capabilities from file %s', fn)
        logger.info('Sheets: %s', book.sheetnames)
        sheetIn = "Capabilities"
        sheetOut = "Parsed_Capabilities"
        if sheetIn in book.sheetnames:
            # Меняем вкладку
            sheet = book[sheetIn]
        else:
            logger.exception('Вкладка %s в таблице не найдена', sheetIn)
            exit(0)
        logger.info('Вкладка %s', sheet.title)
        s = []
        row: object
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value != None:
                    s.append(str([cell.value ]).rstrip("'/]"))
    else:
        logger.error('File type %s undefined', ext)
        exit(1)
else:  # Файл трассировки не найден
    logger.error("File %s not found %s >", os.path.splitext(fn)[0], cwd)
    exit

# В дальнейшем вывод производится в заданный файл вывода
# ( стандартный вывод на экран или в файл на диске)+++ или в файл Excel
if file_output:  # <================================================================================================
    ff = fnO
else:
    ff = sys.stdout

with open(ff, 'w') as fO:
    # вариант без WITH
    # fO = open(fnO, 'w')
    # print("?", file = fO)
    # fO.close()
    ##
    logger.info("Трассировка %s содержит %s строк", fn, len(s))
    print("Трассировка", fn, " содержит ", len(s), " строк", file=fO)
    print("", file=fO)
    ##
    logger.info('Определим, трассировка декодирована TraceViewer-ом из tmf или самой eNB')
    cnt_TMF = 0  # Обнуляем счетчик нахождений паттерна TMF
    cnt_eNB = 0  # Обнуляем счетчик нахождений паттерна eNB
    # Перебираем все строки для сбора бэндов EUTRA
    for i in range(0, len(s)):
        pos = s[i].find(Patt_SBtmf)
        if (pos >= 0):
            cnt_TMF += 1
        else:
            pos = s[i].find(Patt_SBeNB)
            if (pos >= 0):
                cnt_eNB += 1
    corr_eNB = 0  # Коррекция позиции  FGI для tmf не требуется
    corr_eNB2 = 0  # Коррекция позиций ueCapabilityRAT для tmf не требуется
    if cnt_TMF > cnt_eNB:
        print("TMF format decided", "[", cnt_TMF, ":", cnt_eNB, "]", file=fO)
        logger.info("TMF format decided")
        Patt_SB = Patt_SBtmf
        Patt_UtraBands = Patt_UtraBands_tmf
        Patt_GeRANBands = Patt_GeRANBands_tmf
        Patt_FGI8 = Patt_FGI8_tmf
        Patt_FGI9 = Patt_FGI9_tmf
        Patt_FGI10 = Patt_FGI10_tmf
        Patt_geranCS = Patt_geranCS_tmf
        Patt_geranPS = Patt_geranPS_tmf
        Patt_UTRA = Patt_UTRA_tmf
        Patt_UERATcap = Patt_UERATcap_tmf
        Patt_BWclassUL = Patt_BWclassUL_tmf
        Patt_BWclassDL = Patt_BWclassDL_tmf
        Patt_MIMO = Patt_MIMO_tmf
        Patt_CC10 = Patt_CC10_tmf
        Patt_Band10 = Patt_Band10_tmf
        TMF = True
        eNB = False
    else:
        logger.info("eNB format decided")
        print("eNB format decided", "[", cnt_eNB, ":", cnt_TMF, "]", file=fO)
        Patt_SB = Patt_SBeNB
        Patt_UtraBands = Patt_UtraBands_eNB
        Patt_GeRANBands = Patt_GeRANBands_eNB
        Patt_FGI8 = Patt_FGI8_eNB
        Patt_FGI9 = Patt_FGI9_eNB
        Patt_FGI10 = Patt_FGI10_eNB
        Patt_geranCS = Patt_geranCS_eNB
        Patt_geranPS = Patt_geranPS_eNB
        Patt_UTRA = Patt_UTRA_eNB
        Patt_UERATcap = Patt_UERATcap_eNB
        corr_eNB = 1  # Коррекция позиции  FGI = на 1 позицию
        corr_eNB2 = 1  # на 1 позицию правее
        Patt_BWclassUL = Patt_BWclassUL_eNB
        Patt_BWclassDL = Patt_BWclassDL_eNB
        Patt_MIMO = Patt_MIMO_eNB
        Patt_CC10 = Patt_CC10_eNB
        Patt_Band10 = Patt_Band10_eNB
        Patt_StartComb = Patt_StartCombeNB
        Patt_endComb = Patt_endCombeNB
        Patt_UEacc = Patt_UEacc_eNB
        Patt_UEcat = Patt_UEcat_eNB
        eNB = True
        TMF = False
    print("\n", file=fO)

    ##
    logger.info('Перебираем все строки для поиска UE Access Stratum')
    for i in range(0, len(s)):
        pos = s[i].find(Patt_UEacc)
        if (pos >= 0):
            UEaccS.append(EndStrFilter(s[i][pos:]))
    print("Найденные 3gpp Access Stratum:", len(UEaccS), file=fO)
    print("=================================", file=fO)
    for i in range(0, len(UEaccS)):
        print(UEaccS[i], end='', file=fO)
    print("\n", file=fO)

    ##
    logger.info('Перебираем все строки для сбора категорий')
    for i in range(0, len(s)):
        pos = s[i].find(Patt_UEcat)
        if (pos >= 0):
            UEcats.append(EndStrFilter(s[i][pos + 2:]))
    print("Найденные категории LTE:", len(UEcats), file=fO)
    print("=================================", file=fO)

    S_UEcat = []
    S_UEcatUL = []
    S_UEcatDL = []
    for i in range(0, len(UEcats)):
        print(UEcats[i], end='\n', file=fO)
        if 'DL' in UEcats[i]:
            S_UEcatDL.append(int(valfrombrackets(UEcats[i])))
        elif 'UL' in UEcats[i]:
            S_UEcatUL.append(int(valfrombrackets(UEcats[i])))
        else:
            S_UEcat.append(int(valfrombrackets(UEcats[i])))
    print("\n", file=fO)
    S_UEcat.sort()
    S_UEcatDL.sort()
    S_UEcatUL.sort()
    logger.info('LTE Categories: %s', S_UEcat)
    logger.info('LTE DL Categories: %s', S_UEcatDL)
    logger.info('LTE UL Categories: %s', S_UEcatUL)

    ##
    logger.info('Обнуляем начальное значение несущих EUTRA')
    S_EUTRA = [] #  номера LTE bands
    nb = 0
    ##
    logger.info('Перебираем все строки для сбора бэндов EUTRA')
    for i in range(0, len(s)):
        pos = s[i].find(Patt_SB)
        if (pos >= 0):
            Bands.append([0, 0, 0])
            eUtraBands.append("Band  " + (valfrombrackets(s[i][pos + 15:])))
            S_EUTRA.append(int (valfrombrackets (s[i][pos + 15:])))
            Bands[nb][0] = int(valfrombrackets(s[i][pos + 15:]))
            nb += 1
    S_EUTRA.sort()
    logger.info('LTE bands: %s', S_EUTRA)

    NumSBr12 = -1  # инициируем переменную количества бэндов с расширением для R12 значением -1 чтобы она могла быть индексом бэнда
    SBr12Found = False  # инициируем переменную наличия блока расширения для R12 значением False
    # для начала ипринимаем за паттерны UL/DL паттерны для блока V12
    Patt_dl = Patt_dl256
    Patt_ul = Patt_ul64
    for i in range(0, len(s) - 2):
        pos = s[i].find(Patt_SBr12)
        if (pos >= 0):
            SBr12Found = True  # блок модуляций для R12xx присутствует
            NumSBr12 += 1  # количество блоков модуляций для r12 ++
            #print("---SupportedBandEUTRA-v12 =", NumSBr12, file=fO)
            # Если найденный блок v1250, переопределим паттерны для переиспользования кода
            pos = s[i].find(Patt_SBr1250)
            if (pos >= 0):
                Patt_dl = Patt_dl256_1250
                Patt_ul = Patt_ul64_1250
            if TMF:
                Patt_dl = Patt_dl256tmf
                Patt_ul = Patt_ul64tmf
                #   Поиск DL256
            # if ((s[i + 1].find(Patt_dl) >= 0) or (s[i + 2].find(Patt_dl) >= 0)):
            if ((Patt_dl in s[i + 1]) or (Patt_dl in s[i + 2])):
                 eUtraBands[NumSBr12] = eUtraBands[NumSBr12] + " 256QAM DL"
                 Bands[NumSBr12][1] = 8  # bits
#                if (Bands[NumSBr12][0] = PrimaryEUTRABand)
                 if (Bands[NumSBr12][0] in (PrimaryEUTRABand, SecondaryEUTRABands)):
                    S256DL = True #256QAM for DL of some used bands (primary or secondary)
            else:
                eUtraBands[NumSBr12] = eUtraBands[NumSBr12] + " 64QAM  DL"
                Bands[NumSBr12][1] = 6  # bits
            #   Поиск UL64
            if ((s[i + 1].find(Patt_ul) >= 0) or (s[i + 2].find(Patt_ul) >= 0)):
                eUtraBands[NumSBr12] = eUtraBands[NumSBr12] + " 64QAM UL"
                Bands[NumSBr12][2] = 6  # bits
            else:
                eUtraBands[NumSBr12] + eUtraBands[NumSBr12] + " 16QAM UL"
                Bands[NumSBr12][2] = 4  # bits
    print("\n", file=fO)

    if not (SBr12Found):  # блок модуляций для r12 отсутствует
        for j in range(nb):
            eUtraBands[j] = eUtraBands[j] + " 64QAM DL 16QAM UL \n"
            Bands[j][1] = 6  # bits
            Bands[j][2] = 4  # bits


    print("Найденные EUTRA bands ( с расширением до v12.50):", len(eUtraBands), file=fO)
    print("====================================================", file=fO)
    for i in range(0, len(eUtraBands)):
        print(eUtraBands[i], end='\n', file=fO)
    print("\n", file=fO)
    ##
    S_UTRA = []
    logger.info('Перебираем все строки для сбора бэндов UTRA')
    for i in range(0, len(s)):
        pos = s[i].find(Patt_UtraBands)
        if (pos >= 0):
            UtraBands.append(EndStrFilter(s[i][pos + 28:]))
            S_UTRA.append(int(valfrombrackets(s[i][pos + 28:]))+1)
    print("Найденные UTRA bands:", len(UtraBands), file=fO)
    print("========================", file = fO)
    for i in range(0, len(UtraBands)):
        print(UtraBands[i], end='\n', file=fO)
    print("\n", file=fO)
    S_UTRA.sort()
    logger.info('UTRA bands: %s', S_UTRA)

    # Перебираем все строки для сбора бэндов GERAN
    S_GERAN = []
    for i in range(0, len(s)):
        pos = s[i].find(Patt_GeRANBands)
        if (pos >= 0):
            # GeRANBands.append(s[i][pos + 25:])
            GeRANBands.append(valfromGSM(s[i]))
            S_GERAN.append(valfromGSM(s[i]))
    print("Найденные GeRAN bands:", len(GeRANBands), file=fO)
    print("========================", file = fO)
    for i in range(0, len(GeRANBands)):
        print(GeRANBands[i], end='\n', file=fO)
    print("\n", file=fO)
    S_GERAN.sort()
    logger.info('GERAN bands: %s', S_GERAN)

    if fgi_out:
        # Поиск FGI r8
        logger.info('Поиск FGI r9')
        for i in range(0, len(s)):
            pos = s[i].find(Patt_FGI8)
            if (pos >= 0):
                FGI8Txt = s[i][pos + 30 + corr_eNB:pos + 30 + 32 + corr_eNB]
                if len(FGI8Txt) > 0:
                    print("         0               1               ", file=fO)
                    print("         0123456789ABCDEF0123456789ABCDEF", file=fO)
                    print("         +---------------+---------------+", file=fO)
                    print("FGI r8: ", FGI8Txt, file=fO)
                else:
                    print("No FGI r8 found", file=fO)
                    logger.warning("No FGI r8 found")
        print("\n", file=fO)
        ##
        logger.info('Поиск FGI r9')
        for i in range(0, len(s)):
            pos = s[i].find(Patt_FGI9)
            if (pos >= 0):
                FGI9Txt = s[i][pos + 33 + corr_eNB:pos + 33 + 32 + corr_eNB]
                if len(FGI9Txt) > 0:
                    print("         0               1               ", file=fO)
                    print("         0123456789ABCDEF0123456789ABCDEF", file=fO)
                    print("         +---------------+---------------+", file=fO)
                    print("FGI r9: ", FGI9Txt, file=fO)
                else:
                    print("No FGI r9 found", file=fO)
                    logger.warning("No FGI r9 found")
        print("\n", file=fO)
        ##
        logger.info('Поиск FGI r10')
        for i in range(0, len(s)):
            pos = s[i].find(Patt_FGI10)
            if (pos >= 0):
                FGI10Txt = s[i][pos + 32 + corr_eNB:pos + 32 + 32 + corr_eNB]
                if len(FGI10Txt) > 0:
                    print("         0               1               ", file=fO)
                    print("         0123456789ABCDEF0123456789ABCDEF", file=fO)
                    print("         +---------------+---------------+", file=fO)
                    print("FGI r10:", FGI10Txt, file=fO)
                else:
                    print("No FGI r10 found", file=fO)
                    logger.warning("No FGI r10 found")
        print("\n", file=fO)
    ##
    logger.info('Поиск капабилити 2g/3g')
    geranPS_pos = len(s)  # по умолчанию начало блоков GeRAN PS/CS//UTRAN = последней строке
    geranCS_pos = len(s)
    UTRAN_pos = len(s)
    ##
    logger.info('Пытаемся найти истинные начала блоков GeRAN PS/CS//UTRAN')
    for i in range(0, len(s)):
        pos = s[i].find(Patt_geranPS)
        if (pos >= 0):
            geranPS_pos = i
            break
    for i in range(0, len(s)):
        pos = s[i].find(Patt_geranCS)
        if (pos >= 0):
            geranCS_pos = i
            break
    for i in range(0, len(s)):
        pos = s[i].find(Patt_UERATcap)
        if (pos >= 0):
            UTRAN_pos = i
            break
    if geranPS_pos == len(s):
        print("geranPS block is absent", file=fO)
        logger.error("geranPS block is absent")
    if geranCS_pos == len(s):
        print("geranCS block is absent", file=fO)
        logger.error("geranCS block is absent")
    if UTRAN_pos == len(s):
        print("UTRAN   block is absent", file=fO)
        logger.error("UTRAN   block is absent")
    # упорядочиваем блоки
    smin = min(geranPS_pos, geranCS_pos, UTRAN_pos)
    smax = max(geranPS_pos, geranCS_pos, UTRAN_pos)
    ##
    logger.info('Ищем строку капабилити в блоке UTRAN')
    if utran_out and UTRAN_pos < len(s):
        if UTRAN_pos == smax:
            k = len(s)
            print("<UTRAN block is 3rd>", file=fO)
        elif UTRAN_pos in range(smin + 1, smax):
            k = smax
            print("<UTRAN block is 2nd>", file=fO)
        else:  # UTRAN_pos == smin
            if geranPS_pos > geranCS_pos:
                k = geranCS_pos
                print("<UTRAN block is 1st, geranCS is 2nd>", file=fO)
            else:
                k = geranPS_pos
                print("<UTRAN block is 1st, geranPS is 2nd>", file=fO)
        for i in range(UTRAN_pos, k):
            pos2 = s[i].find(Patt_UERATcap)  # позиция начала паттерна в строке
            if (pos2 >= 0):
                pos3 = s[i].find(" (")  # позиция конца паттерна в строке
                if pos3 == None:
                    UTRANcapTxt = s[i][pos + 19 + corr_eNB2:]
                else:
                    UTRANcapTxt = s[i][pos + 19 + corr_eNB2:pos3]
        print("UTRAN capabilities: \n =0x", UTRANcapTxt, sep='', file=fO)
        if utrangeranbinary:
            bits = 4 * len(UTRANcapTxt)
            if (len(UTRANcapTxt) > 1):
                UTRANcapBytes = bytes.fromhex(UTRANcapTxt)
                print("\nBinary = ", end='', file=fO)
                UTRANcapBits = []
                n = int.from_bytes(UTRANcapBytes, byteorder='big', signed=False)
                for i in (range(bits - 1, 0, -1)):
                    if ((2 ** i & n) > 0):
                        UTRANcapBits.append("1")
                    else:
                        UTRANcapBits.append("0")
                for i in range(0, bits - 1):
                    print(UTRANcapBits[i], file=fO)
                print("\n", file=fO)
        print("\n", file=fO)
    ##
    logger.info('Поиск капабилити для GeranPS')
    if geran_out and (geranPS_pos < len(s)):
        if geranPS_pos == smax:
            k = len(s)
            print("<geranPS block is 3rd>", file=fO)
        elif geranPS_pos in range(smin + 1, smax):
            k = smax
            print("<geranPS block is 2nd>", file=fO)
        else:  # geranPS_pos == smin
            if geranCS_pos > UTRAN_pos:
                k = UTRAN_pos
                print("<geranPS block is 1st, UTRAN is 2nd>", file=fO)
            else:
                k = geranCS_pos
                print("<geranPS block is 1st, geranCS is 2nd>", file=fO)
        ##
        logger.info('Ищем строку капабилити в блоке geranPS')
        for i in range(geranPS_pos, k):
            pos2 = s[i].find(Patt_UERATcap)  # позиция начала паттерна в строке
            if (pos2 >= 0):
                pos3 = s[i].find(" (")
                if pos3 == None:
                    GeranPScapTxt = s[i][pos + 19 + corr_eNB2:]
                else:
                    GeranPScapTxt = s[i][pos + 19 + corr_eNB2:pos3]
        if (len(GeranPScapTxt) > 1):
            print("\nGERAN PS capabilities: \n =0x", GeranPScapTxt, sep='', file=fO)
            if utrangeranbinary:
                bits = 4 * len(GeranPScapTxt)
                GeranPScapBytes = bytes.fromhex(GeranPScapTxt)
                if (len(GeranPScapTxt) > 1):
                    print("\nGERAN PS capabilities: \n =0x", GeranPScapTxt, sep='', file=fO)
                    ## print(GeranPScapBytes,file = fO)
                    GeranPScapBits = []
                    n = int.from_bytes(GeranPScapBytes, byteorder='big', signed=False)
                    for i in (range(bits - 1, 0, -1)):
                        if ((2 ** i & n) > 0):
                            GeranPScapBits.append("1")
                        else:
                            GeranPScapBits.append("0")
                    print("\nBinary = ", end='', file=fO)
                    for i in range(0, bits - 1):
                        print(GeranPScapBits[i], end='', file=fO)
                    print("\n", file=fO)
        print("\n", file=fO)

    logger.info('Поиск капабилити для GeranCS')
    if geran_out and geranCS_pos < len(s):
        if geranCS_pos == smax:
            k = len(s)
            print("<geranCS block is 3rd>", file=fO)
        elif geranCS_pos in range(smin + 1, smax):
            k = smax
            print("<geranCS block is 2nd>", file=fO)
        else:  # geranCS_pos == smin
            if geranPS_pos > UTRAN_pos:
                k = UTRAN_pos
                print("<geranCS block is 1st, UTRAN is 2nd>", file=fO)
            else:
                k = geranPS_pos
                print("<geranCS block is 1st, geranPS is 2nd>", file=fO)
            ##
        logger.info('Ищем строку капабилити в блоке geranCS')
        for i in range(geranCS_pos, k):
            pos2 = s[i].find(Patt_UERATcap)  # позиция начала паттерна в строке
            if (pos2 >= 0):
                pos3 = s[i].find(" (")  # позиция конца паттерна в строке
                if pos3 == None:
                    GeranCScapTxt = s[i][pos + 19 + corr_eNB2:]
                else:
                    GeranCScapTxt = s[i][pos + 19 + corr_eNB2:pos3]
        if (len(GeranCScapTxt) > 1):
            print("GERAN CS capabilities: \n =0x", GeranCScapTxt, sep='', file=fO)
            ## print(GeranCScapBytes,file = fO)
            GeranCScapBits = []
            bits = 4 * len(GeranCScapTxt)
            if utrangeranbinary:
                GeranCScapBytes = bytes.fromhex(GeranCScapTxt)
                n = int.from_bytes(GeranCScapBytes, byteorder='big', signed=False)
                for i in (range(bits - 1, 0, -1)):
                    if ((2 ** i & n) > 0):
                        GeranCScapBits.append("1")
                    else:
                        GeranCScapBits.append("0")
                print("\nBinary = ", end='', file=fO)
                for i in range(0, bits - 1):
                    print(GeranCScapBits[i], end='', file=fO)
                print("\n", file=fO)
    print("\n", file=fO)

    # Костыль для поиска UL 256QAM для R14
    CCr14Mod = []  # Массив/список для хранения расширений UL 256QAM для R14
    NumCCr14 = 0  # инициируем переменную количества комбинаций с UL 256QAM для R14
    CCr14Found = False  # инициируем переменную наличия блока расширения для R14 значением False

    for i in range(0, len(s)):
        pos = s[i].find(Patt_ul256r14)
        if (pos >= 0):
            CCr14Found = True  # блок модуляций для r14 присутствует
            NumCCr14 += 1  # количество блоков модуляций для r14 ++
            #print("---SupportedCCEUTRA-r14 =", NumCCr14, file=fO)
            if 'supported' in s[i]:
                CCr14Mod.append(8)  # bits
            else:
                CCr14Mod.append(None)
    print("\n", file=fO)
    if CCr14Found:
        logger.info('Найдена поддержка компонент с UL256QAM r14.3')

    logger.info('Поиск начала блока комбинаций несущих r10')
    i = 0
    while (i < len(s) and s[i].find(Patt_StartComb) < 0):
        i += 1
    Comb_Start = i
    logger.info('Поиск конца блока комбинаций несущих r10')
    while (i < len(s) and s[i].find(Patt_endComb) < 0):
        i += 1
    Comb_End = i
    if (Comb_End > Comb_Start):
        logger.info('Перебираем ВСЕ строки блока комбинаций несущих')
        NCCs = 0
        Ncarr = 0
        for i in range(Comb_Start, Comb_End + 1):
            if (s[i].find(Patt_bc) >= 0):
                NCCs += 1
            endcc = True
            while (endcc):
                if (s[i].find(Patt_CC10) >= 0):
                    CCs.append([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])  # Добавляем пустую запись для первой несущей в найденной комбинации несущих
                    Ncarr += 1  # 0,1,2,3,4,5,6,7,8,9,A - поля
                    # cбор и заполнение параметров для найденной несущей
                    CCs[Ncarr - 1][0] = NCCs
                    CCs[Ncarr - 1][1] = int(valfrombrackets(s[i][pos + 20:]))
                    # находим параметры модуляций UL/DL для текущего Band  в  списке Bands  и копируем их в поля 1..3
                    for j in range(0, len(eUtraBands)):
                        if (Bands[j][0] == CCs[Ncarr - 1][1]):
                            CCs[Ncarr - 1][2] = Bands[j][1]
                            CCs[Ncarr - 1][3] = Bands[j][2]

                            # contigous carriers in UL -->[5]
                if s[i].find(Patt_BWclassUL) >= 0:
                    temp_int = int(valfrombrackets(s[i][pos + 30:]))
                    CCs[Ncarr - 1][5] = temp_int
                    if (temp_int == 1):  # class "B"
                        CCs[Ncarr - 1][5] = 2
                        CCs[Ncarr - 1][8] = 10
                    elif (temp_int == 0):  # class "A"
                        CCs[Ncarr - 1][5] = 1
                        CCs[Ncarr - 1][8] = 20
                    else:  # classes {C,D...Z}
                        CCs[Ncarr - 1][5] = temp_int
                        CCs[Ncarr - 1][8] = 20

                    # contigous carriers in DL --> [4]
                if s[i].find(Patt_BWclassDL) >= 0:
                    temp_int = int(valfrombrackets(s[i][pos + 30:]))
                    if (temp_int == 1):  # class "B"
                        CCs[Ncarr - 1][4] = temp_int + 1
                        CCs[Ncarr - 1][7] = 10
                    elif (temp_int == 0):  # class "A"
                        CCs[Ncarr - 1][4] = temp_int + 1
                        CCs[Ncarr - 1][7] = 20
                    else:  # classes {C,D...Z}
                        CCs[Ncarr - 1][4] = temp_int
                        CCs[Ncarr - 1][7] = 20
                if s[i].find(Patt_MIMO) >= 0:
                    temp_int = int(valfrombrackets(s[i][pos + 45:]))
                    if (temp_int == 0):
                        CCs[Ncarr - 1][6] = 2
                    elif (temp_int == 1):
                        CCs[Ncarr - 1][6] = 4 # 4 layers
                        if (CCs[Ncarr-1][1] in (PrimaryEUTRABand,SecondaryEUTRABands)):
                            S_4x4 = True
                    elif (temp_int == 2):
                        CCs[Ncarr - 1][6] = 8 # 8 layers
                        if (CCs[Ncarr - 1][2] in (PrimaryEUTRABand, SecondaryEUTRABands)):
                            S_8x8 = True  # 8 layers for some of use band(s)
                    if (CCs[Ncarr - 1][1] > 32) & (CCs[Ncarr - 1][1] < 49):  # TDD ULDL_config#3 (DL/UL=6/3): DL = 0.6*0.625=0.375 | UL = 0.3*0.625 = 0.1875
                        IsFDD = False
                        CCs[Ncarr - 1][9] = int(
                            0.375 * CCs[Ncarr - 1][2] * CCs[Ncarr - 1][4] * CCs[Ncarr - 1][6] * CCs[Ncarr - 1][7])
                        CCs[Ncarr - 1][10] = int(0.1875 * CCs[Ncarr - 1][3] * CCs[Ncarr - 1][5] * CCs[Ncarr - 1][8])
                    else:  # FDD
                        IsFDD = True
                        CCs[Ncarr - 1][9] = int(
                            0.625 * CCs[Ncarr - 1][2] * CCs[Ncarr - 1][4] * CCs[Ncarr - 1][6] * CCs[Ncarr - 1][7])
                        CCs[Ncarr - 1][10] = int(0.625 * CCs[Ncarr - 1][3] * CCs[Ncarr - 1][5] * CCs[Ncarr - 1][8])
                endcc = False
        #       print("\n",file = fO)
        print("Найдено ", NCCs, " комбинаций несущих:", file=fO)
        #       print("\n",file = fO)
        if NCCs > 0:
            nccUL_r14=0 # обнуляем номер UL компоненты для добавки данных UL модуляций R1430
            #Вносим коррекции в UL при наличии блока UL 256QAM R14 и включенного свитча обработки R14
            if (R14_enabled and CCr14Found):
                logger.info('Обрабатываем расширение модуляций для R14')
                for l in range(Ncarr):
                    # если у текущей компоненты UL активен (имеет не менее 1 пространственного канала)
                    # и для этой комбинации разрешен 256QAM UL (не None)
                    if (CCs[l][5]>0 and CCr14Mod[nccUL_r14]>0):
                        CCs[l][3] = CCr14Mod[nccUL_r14]
                        CCs[l][10] = int(0.625 * CCs[l][3] * CCs[l][5] * CCs[l][8])
                        nccUL_r14 += 1
                        if CCs[l][2] in (PrimaryEUTRABand, SecondaryEUTRABands):
                            S256UL = True
            if CCr14Found:
                logger.info('Найдено %s компонент с UL 256QAM',nccUL_r14)
            ##

            logger.info('Создаем список комбинаций ')
            for i in range(NCCs):  # комбинации
                # заполняем список комбинаций суммами компонентных несущих
                CCsCC.append([i + 1, 0, 0])
                for j in range(Ncarr):  # компоненты
                    if (CCs[j][0] == i + 1):
                        CCsCC[i][1] = CCsCC[i][1] + CCs[j][9]
                        CCsCC[i][2] = CCsCC[i][2] + CCs[j][10]
            if table_format:
                t1.field_names = ["№комб", "DL Mbps", "UL_Mbps"]
                #       Alignments
                t1.align["№комб"] = "r"
                t1.align["DL Mbps"] = "r"
                t1.align["UL_Mbps"] = "r"
            else:
                print("№комб.", " ", "DL Mbps", "      ", "UL_Mbps", file=fO)
            if not (table_format):
                print("================================", file=fO)
            ##
            logger.info('Вывод всех комбинаций')
            for i in range(NCCs):
                if not (table_format):
                    print(CCsCC[i][0], "\t", CCsCC[i][1], "\t\t", CCsCC[i][2], end="\n", file=fO)
                else:
                    t1.add_row([CCsCC[i][0], CCsCC[i][1], CCsCC[i][2]])
            if table_format:
                print(t1, file=fO)
        print("\n", file=fO)
        #        print("\n",file = fO)

        ##
        logger.info('Вывод списка компонентных несущих')
        if Ncarr > 0:
            print("Найдено ", Ncarr, " компонентов несущих:", file=fO)
            print("Для TDD расчет ожидаемой скорости выполняется для конфигурации TDD #3 (6DL/3UL таймслотов)", file=fO)
            #            print("\n",file = fO)
            if not (table_format):
                print("СС#     №комб. Band#   DL_бит  UL_бит  №DL     №UL    MIMO    DLbwMHz  ULbwMHz DL,Mbps UL.Mbps",
                      file=fO)
                print(
                    "=================================================================================================",
                    file=fO)
            else:
                t2.field_names = ["№комб", "СС#", "Band#", "DL_bits", "UL_bits", "№DL", "№UL", "MIMO", "DLbwMHz",
                                  "ULbwMHz", "DL,Mbps", "UL.Mbps"]
            #           t2.align = ["r","r","r","r","r","r","r","r","r","r","r","r"]

            prevComb = 1
            for i in range(Ncarr):  # компоненты 0...Ncarr-1
                # Ищем S_CA7C
                if ((CCs[i][1] == 7) and (CCs[i][4] == 2)):
                    S_CA7C = True
                if CCs[i][0] > prevComb:  # если текущая комбинация увеличилась
                    temp_int = CCs[i][0] - 1  # комбинация = прошлая, выведем для нее суммарные скорости:
                    if table_format:
                        t2.add_row([temp_int, "SUM", "--", "-", "-", "-", "-", "-", "--", "--", CCsCC[temp_int - 1][1],
                                    CCsCC[temp_int - 1][2]])
                        t2.add_row(["", "", "", "", "", "", "", "", "", "", "", ""])
                    else:
                        if temp_int < 10:  # Alignment (not Table)
                            print("Комб.==", temp_int, "=" * 69, CCsCC[temp_int - 1][1], "===", CCsCC[temp_int - 1][2],
                                  file=fO)
                        else:
                            print("Комб.==", temp_int, "=" * 68, CCsCC[temp_int - 1][1], "===", CCsCC[temp_int - 1][2],
                                  file=fO)
                        print("\n", file=fO)
                prevComb = CCs[i][0]  # сохраняем номер текущей комбинации как предыдущий
                if not (table_format):
                    print(i + 1, "\t", end='', file=fO)  # печать индекса = номер CC компоненты
                    for j in range(11):  # колонки 0..10
                        print(CCs[i][j], "\t", end='', file=fO)  # печать колонок 0..10
                    print("\n", file=fO)
                else:
                    t2.add_row(
                        [CCs[i][0], i + 1, CCs[i][1], CCs[i][2], CCs[i][3], CCs[i][4], CCs[i][5], CCs[i][6], CCs[i][7],
                         CCs[i][8], CCs[i][9], CCs[i][10]])
            temp_int = NCCs - 1  # комбинация = последняя, выведем для нее суммарные скорости:
            if not (table_format):
                if temp_int + 1 < 10:  # Alignment (not Table)
                    print("Комб.==", temp_int + 1, "=" * 69, CCsCC[temp_int][1], "===", CCsCC[temp_int][2], file=fO)
                else:
                    print("Комб.==", temp_int + 1, "=" * 68, CCsCC[temp_int][1], "===", CCsCC[temp_int][2], file=fO)
            else:
                t2.add_row([temp_int + 1, "SUM", "--", "-", "-", "-", "-", "-", "--", "--", CCsCC[temp_int][1],
                            CCsCC[temp_int][2]])
            if table_format:
                print(t2, file=fO)
            print("\n", file=fO)
            logger.info("UE Capabilities parsing finished")
            if Excel_output:
                logger = logging.getLogger('XLSproc')
                if "Parsed_Capabilities" not in book.sheetnames:
                    # Add new sheet, fill it and save book copy with new sheet
                    ws1 = book.create_sheet("Parsed_Capabilities")  # insert sheet at the end (by default)
                    logging.info('Вкладка Cap.Info создана')
                else:
                    logging.info('Вкладка Cap.Info уже существует')
                sheet = book["Parsed_Capabilities"]
                sheet.sheet_properties.tabColor = "1072BA"
                sheet['A2'] = 'LTE category DL+UL'
                sheet['B2'] = str(S_UEcat)
                sheet['A3'] = 'LTE category DL'
                sheet['B3'] = str(S_UEcatDL)
                sheet['A4'] = 'LTE category UL'
                sheet['B4'] = str(S_UEcatUL)
                sheet['A5'] = 'EUTRA_Bands'
                sheet['B5'] = str(S_EUTRA)
                sheet['A6'] = 'UTRA_Bands'
                sheet['B6'] = str(S_UTRA)
                sheet['A7'] = 'GeRAN_Bands'
                sheet['B7'] = str(S_GERAN)
                sheet['A8'] = '256QAM for used bands'
                if S256DL and S256UL:
                    sheet['B8'] = 'DL + UL'
                else:
                    if S256DL:
                        sheet['B8'] = 'DL'
                    elif S256UL:
                        sheet['B8'] = 'UL'
                    else:
                        sheet['B8'] = ''
                currentcell = sheet['B8']
                currentcell.font = Font(size=12, underline='none', color=colors.DARKBLUE, bold=True, italic=False)
                sheet['A9'] = 'MIMO 4x4 for used bands'
                if S_4x4:
                    sheet['B9'] = 'Supported'
                else:
                    sheet['B9'] = 'Not Supported'
                currentcell = sheet['B9']
                currentcell.font = Font(size=12, underline='none', color=colors.DARKBLUE, bold=True, italic=False)
                sheet['A10'] = 'CA 7c, 256 QAM, MIMO 4x4'
                S_Qualcomm = (((S256DL or S256UL) and S_4x4) and S_CA7C)
                if S_Qualcomm:
                    sheet['B10'] = 'Supported'
                else:
                    sheet['B10'] = 'Not Supported'
                currentcell = sheet['B10']
                currentcell.font = Font(size=12, underline='none', color=colors.DARKBLUE, bold=True, italic=False)
                ij = 15
                sheet.cell(row=ij, column=1, value='UE access stratum : '+str(UEaccS))
                ij+=1
                sheet.cell(row=ij, column=1, value='UE categories : '+str(UEcats))
                ij+=1
                sheet.cell(row=ij, column=1, value='FGI r8 = '+str(FGI8Txt))
                ij+=1
                sheet.cell(row=ij, column=1, value='FGIr9 = '+str(FGI9Txt))
                ij+=1
                sheet.cell(row=ij, column=1, value='FGIr10 = '+str(FGI10Txt))
                ij+=1
                sheet.cell(row=ij, column=1, value='UTRAN capabilities : '+str(UTRANcapTxt))
                ij+=1
                sheet.cell(row=ij, column=1, value='GERAN CS capabilities : '+str(GeranCScapTxt))
                ij+=1
                sheet.cell(row=ij, column=1, value='GERAN PS capabilities : '+str(GeranPScapTxt))
                ij+=1
                sheet.cell(row=ij, column=1, value='CA Combinations : ')
                currentcell = sheet.cell(row=ij,column=1)
                currentcell.font = Font(size=12, underline='none', color=colors.DARKBLUE, bold=True, italic=False)
                ij+=1
                Menu = ['Comb#', 'Band', 'DLbits', 'ULbits', 'DL#', 'UL#', 'MIMO', 'DL bandwith', 'UL bandwith',
                 'DL Throughput', 'UL Throughput']
                for i in range (11):
                    sheet.cell(row=ij, column=i+1, value= Menu[i])
                    currentcell = sheet.cell(row=ij,column=i+1)
                    currentcell.font = Font(size=12, underline='none', color=colors.YELLOW, bold=True, italic=False)
                    currentcell.fill = PatternFill(fill_type='solid', start_color=colors.DARKBLUE,
                                                     end_color=colors.DARKBLUE)
                ij+=1
                for i in range(Ncarr):
                    for j in range(11):
                        sheet.cell(row=ij+i, column=j+1, value=CCs[i][j])
                        # currentcell = sheet.cell(row=ij+i, column=j+1)
                        # currentcell.font = Font(size=10, underline='none', color=colors.DARKBLUE, bold=False, italic=True)
                        sheet.cell(row=ij + i, column=j + 1).font = Font(size=10, underline='none', color=colors.DARKBLUE, bold=False, italic=True)
                logging.info('Вкладка Cap.Info заполнена')
                book.save(fnX)
                # book.save('fnX')
                book.save('LatestParsed'+ext)
                logging.info('Копия файла сохранена в файле %s', fnX)
    else:
        print("CC combinations not found", file=fO)
        logger.warning("CC combinations not found")
