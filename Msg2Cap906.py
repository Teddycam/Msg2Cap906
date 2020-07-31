from Patterns import *
from tkinter import *
# from tkinter import font
import os
import sys
import openpyxl
import logging
import tkinter.filedialog
from prettytable import PrettyTable
from openpyxl.styles import Font, Fill  # Подключаем стили для текста
from openpyxl.styles import colors  # Подключаем цвета для текста и ячеек
from openpyxl.styles import PatternFill  # Подключаем стили для ячеек

# Switches
# ========
fgi_out = True
geran_out = True
utran_out = True
utrangeranbinary = True
R14_enabled = True

txt_output = False
table_format = True
Excel_output = False
BandsFilter = True

# FsegoeUI8 = tkFont.Font(family='Segoe UI', size=8, weight='bold')
FsegoeUI8 = Font('Segoe UI', 8, 'bold')

# VARs
TMF = False
eNB = False
S256DL = False # 256QAM DL for some of use band(s)
S256UL = False # 256QAM UL for some of use band(s)
S_4x4 = False # 4 layers for some of use band(s)
S_8x8 = False # 8 layers for some of use band(s)
S_CA7C = False # CA 7C support (CA of 2 carriers in Band 7)
S_Qualcomm = False # Qualcomm's requirement statement value

# internal DB arrays:
Bands = []
CCs = []
CCsCC = []

# Temporary variables
temp_int = 0
temp_str = ""

# Output lists
UEaccS = []
UEcats = []
eUtraBands = []
UtraBands = []
GeRANBands = []
Combinations = []
GeranCScapTxt = ""
GeranPScapTxt = ""
UTRANcapTxt = ""
FGI8txt = ""
FGI9txt = ""
FGI10txt = ""

# Настройка  модуля Logging
formatter = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
logging.basicConfig(level=logging.INFO, format=formatter)
logger = logging.getLogger('_Parser')
logger.setLevel(logging.INFO)

# извлечение числа из строчного значения в скобках вида (25)
def valfrombrackets(inputstr):
    """
    This for parse something
    :param inputstr: this is regular expression
    :return: string value without brackets
    e.g. inputstr = '(bla vla)'
    return = 'bla vla'
    """
    pos01 = inputstr.find("(")
    pos02 = inputstr.find(")")
    if (pos01 >= 0 and pos02 > 0):
        return inputstr[pos01 + 1:pos02]
    else:
        logger.error('ValFromBrackets cannot find any value in %s:', inputstr)
        logger.error('Returned None')
        return None

def valFGIfromquotes(inputstr):
    """
    This for parse something
    :param inputstr: this is regular expression
    :return: string value without quotes
    e.g. inputstr = "'bla vla'"
    return = 'bla vla'
    """
    pos01 = inputstr.find("\'")
    if pos01 <0:
        pos01 = 0
    pos02 = inputstr[pos01+1:].find("\'")
    if (pos01 >= 0 and pos02 == 32):
        return inputstr[pos01 + 1:pos01+33]
    else:
        logger.error('ValFromQuotes cannot find any value in %s:', inputstr)
        logger.error('Returned None')
        return ''


def valfromGSM(inputstr):
    """
    This for parse something
    :param inputstr: this is regular expression
    :return: INTEGER value of gsm frequency
    e.g. inputstr = 'SupportedBandGERAN :  ---- gsm900E(7) ---- 00111***'
    return = 900
    """
    aa = ''
    pos01 = inputstr.find("gsm")
    pos02 = inputstr.find("(")
    if (pos01 >= 0 and pos02 > 0):
        a = inputstr[pos01+3:pos02]
        for b in range(0, len(a)):
            if (a[b] in '0123456789'): aa= aa+a[b]
        if len(a)>0:
            return int(aa)
        else:
            return None
    else:
        logger.error('ValFromGSM cannot find any value in %s:', inputstr)
        logger.error('Returned None')
        return None

def EndStrFilter(inputstr):
    """
    This for parse something
    :param inputstr: this is regular expression
    :return: inputstr up to ')' symbol without /n
    e.g. inputstr = 'ue-Category --- 0x4(4)']'
    return = 'ue-Category --- 0x4(4)'
    """
    pos01 = inputstr.find(")")
    if (pos01 > 0):
        a = inputstr[:pos01+1]
        a = a.rstrip('\n')
        if len(a)>0:
            return a
        else:
            return ''
    else:
        logger.error('EndStrFilter() cannot find any value in %s:', inputstr)
        logger.error('Returned None')
        return ''

def  FillCellSum(r,c,v,ws,clr):
    """
    This for fill cell [r][c] of worksheet ws with style for SUM  
    :param ws: worksheet
            r: raw
            c: column
            v: Value for cell
            clr: color of fill
    :return: filled worksheet 
    e.g. FillCellSum (RowInSheet+i, j+1, SumComb[j], sheet)
    return = worksheet with with filled cells
    """
    ws.cell(row=r, column=c, value=v)
    ws.cell(row=r, column=c).font = Font(name=FSimple,size=12, underline='single', color=DARKRED, bold=True, italic=True)
    ws.cell(row=r, column=c).fill = PatternFill(fill_type='solid', start_color=clrCC , end_color=clrCC)
    return ws

def  FillCellNorm(r,c,v,ws):
    """
    This for fill cell [r][c] of worksheet ws with style for Normal output
    :param ws: worksheet
            r: raw
            c: column
            v: Value for cell
    :return: filled worksheet
    e.g. FillCellNorm(RowInSheet+1, j + 1, CCs[i][j], sheet)
    return = worksheet with filled cells
    """
    ws.cell(row=r, column=c, value=v)
    ws.cell(row=r, column=c).font = Font(name=FSimple,size=10, underline='none', color=DARKBLUE, bold=False, italic=True)
    return ws

def CollectPatt(Slist, Patt, corr):
    """
    This for collect founded in the Slist Patt values to Flist
    :param  Slist: input list of Str
            Patt: search pattern
            corr: correction value for position
    :return: Flist: list of collected values
    e.g. CollectPatt(s, Patt_UEacc)
    return = list of collected values
    """
    Flist = []
    for i in range(0, len(Slist)):
        pos = s[i].find(Patt)
        if (pos >= 0):
            Flist.append(EndStrFilter(s[i][pos + corr:]).replace('-',''))
    return Flist

def CheckCCCinUsed(cc, UPat):
    """
    This for check, is cc in Used  bands ccs Patterns or not
    :param  cc   : string of ccs in combination like '3172'  for compare to possible used patterns
            Upat : array of string wuth used patterns of bands and cc# like ['3172','3272201]
    :return: Boolean (is ccs in Used Pat or not)
    e.g. CCCinUsed = CheckCCCinUsed(CCC[i], UsedPat)
    """
    cciu = False
    for i5 in UPat:
        cciu = cciu or (cc == i5)
    return cciu

def Conv2Bits(Captxt):
    """
    This for convert Hex string with  capabilities to the binary string
    :param  Captxt: Hex string
    :return: CapBits: Binary string
    e.g. GeranPScapBits = Conv2Bits(GeranPSCaptxt)
    return = list of collected values
    """
    bits = 4 * len(Captxt)
    CapBytes = bytes.fromhex(Captxt)
    if (len(Captxt) > 1):
        CapBits = []
        n = int.from_bytes(CapBytes, byteorder='big', signed=False)
        for i in (range(bits - 1, 0, -1)):
            if ((2 ** i & n) > 0):
                CapBits.append("1")
            else:
                CapBits.append("0")
    return CapBits

def GUI(tt,ttX):
    """
    This is GUI for getting switches values and checking output file name and path
    :param tt: output file name
    : return: ttt: [possibly] changed output text file name
              tttX:[possibly] changed output XLS file name
              vv: values list of the switches snd radiobuttons
              ls: list of selected elements of list (not used yet)
    """
    def p1():
        global vv
        vv = []
        global ls
        ls = []
        logger.info('Button1 pressed')
        global ttt
        global tttX
        ttt = textT.get('1.0', END)
        tttX = textX.get('1.0', END)
        logger.info('TXT File name changed to: %s', ttt)
        logger.info('XLS File name changed to: %s', tttX)
        vv.append(var1.get())
        vv.append(var2.get())
        vv.append(var3.get())
        vv.append(var4.get())
        vv.append(var5.get())
        vv.append(var6.get())
        vv.append(var7.get())
        vv.append(var8.get())
        vv.append(var9.get())
        vv.append(var10.get())
        vv.append(var11.get())
        logger.info('Switches: %s', vv)
        ls = listbox1.curselection()
        logger.info('List: %s', ls)
        window.destroy()

    def p2():
        logger.warning('Button2 pressed, cancel')
        window.destroy()
        exit(0)

    def c0():
        logger.info('Radio = %s',var11.get())

        if var11.get()==3: # TXT+XLS
            frameT.pack(side='top', expand = TRUE)
            textT.pack(side='top', fill='both')
            frameX.pack(side='top', expand = TRUE)
            textX.pack(side='top', fill='both')

        elif var11.get()==2: # TXT only
            frameT.pack(side='top', expand=TRUE)
            textT.pack(side='top', fill='both')
            # frameX.pack_forget()
            textX.pack_forget()

        elif var11.get()==1: # Screen only, no files
            # frameT.pack_forget()
            textT.pack_forget()
            # frameX.pack_forget()
            textX.pack_forget()

        else:
            logger.error("No output choosen")
            window.destroy()
            exit(0)

    def c1():
        logger.info('Radio = %s',var11.get())
        var7.set(0)
        textT.pack_forget()
        textX.pack_forget()

    def c2():
        logger.info('Radio = %s',var11.get())
        var7.set(0)
        frameT.pack(side='top', expand=TRUE)
        textT.pack(side='top', fill='both')
        textX.pack_forget()

    def c3():
        logger.info('Radio = %s',var11.get())
        var7.set(1)
        frameT.pack(side='top', expand = TRUE)
        textT.pack(side='top', fill='both')
        frameX.pack(side='top', expand = TRUE)
        textX.pack(side='top', fill='both')


    def ppp(event):
        logger.info('Button1 pressed')

    def eee(event):
        logger.warning('Button2 pressed, cancel')
        exit(0)

    logger = logging.getLogger('_GUI')

    window = Tk()
    var1 = IntVar()
    var2 = IntVar()
    var3 = IntVar()
    var4 = IntVar()
    var5 = IntVar()
    var6 = IntVar()
    var7 = IntVar()
    var8 = IntVar()
    var9 = IntVar()
    var10 = IntVar()
    var11 = IntVar()
    defFont = ('Consolas', 9, 'bold')
    #FsegoeUI8 #'Consolas 9' || 'Segoe 9'//'Segoe UI'

    if '.txt' in ext:
        switches = DEF_TXT_SW
    elif 'xls' in ext:
        switches = DEF_XLS_SW
    else:
        switches = DEF_SCR_SW
    var1.set(switches[0])
    var2.set(switches[1])
    var3.set(switches[2])
    var4.set(switches[3])
    var5.set(switches[4])
    var6.set(switches[5])
    var7.set(switches[6])
    var8.set(switches[7])
    var9.set(switches[8])
    var10.set(switches[9])
    var11.set(switches[10])
    window.title("Msg2Cap")

    frameO = LabelFrame(window,  bd = 2, fg = 'blue', text = 'Output:') # background = 'white',

    frameT = LabelFrame(frameO,  bd = 2, fg = 'blue', text = 'TXT file name:') # background = 'white',
    textT = Text(frameT, height = 2, width = len(tt)+5, bd = 0, font = defFont, wrap = WORD, highlightcolor = 'yellow')
    textT.insert(1.0, tt)

    frameX = LabelFrame(frameO, bd = 2, fg = 'blue', text = 'XLS file name:') # background = 'white',
    textX = Text(frameX, height = 2, width = len(tt)+5, bd = 0, font = defFont, wrap = WORD, highlightbackground = 'yellow')
    textX.insert(1.0, ttX)

    frameLb = LabelFrame(window, bd=2, fg='blue', text='Options:')  # background = 'white',
    listbox1 = Listbox(frameLb, height = 8, width = 15, selectmode = EXTENDED, font = defFont )
    list1 = ['Option №0','Option №1','Option №2','Option №3','Option №4','Option №5','Option №6','Option №7','Option №8',
             'Option №9', 'Option №10','Option №11','Option №12','Option №13','Option №14','Option №15']
    for i in list1: listbox1.insert(END,i)

    frameCB = LabelFrame(window, bd = 2, fg = 'blue', text = 'Parsing options:') # , background = 'white'
    check1 = Checkbutton(frameCB, text = SwitchesNames[0], font = defFont, variable = var1, onvalue = 1, offvalue = 0 ,command = c0)
    check2 = Checkbutton(frameCB, text = SwitchesNames[1], font = defFont, variable = var2, onvalue = 1, offvalue = 0 ,command = c0)
    check3 = Checkbutton(frameCB, text = SwitchesNames[2], font = defFont, variable = var3, onvalue = 1, offvalue = 0 ,command = c0)
    check4 = Checkbutton(frameCB, text = SwitchesNames[3], font = defFont, variable = var4, onvalue = 1, offvalue = 0 ,command = c0)
    check5 = Checkbutton(frameCB, text = SwitchesNames[4], font = defFont, variable = var5, onvalue = 1, offvalue = 0 ,command = c0)
    check6 = Checkbutton(frameCB, text = SwitchesNames[5], font = defFont, variable = var6, onvalue = 1, offvalue = 0 ,command = c0)
    check7 = Checkbutton(frameCB, text = SwitchesNames[6], font = defFont, variable = var7, onvalue = 1, offvalue = 0 ,command = c0)
    check8 = Checkbutton(frameCB, text = SwitchesNames[7], font = defFont, variable = var8, onvalue = 1, offvalue = 0 ,command = c0)
    check9 = Checkbutton(frameCB, text = SwitchesNames[8], font = defFont, variable = var9, onvalue = 1, offvalue = 0 ,command = c0)
    check10 = Checkbutton(frameCB, text = SwitchesNames[9],font = defFont, variable = var10, onvalue = 1, offvalue = 0 ,command = c0)

    frameRB = LabelFrame(window,  bd = 2, fg = 'blue', text = 'Output:', width = len(tt)+5) # background = 'white',
    rbutton1 = Radiobutton( frameRB, text = ' Screen ', overrelief = RAISED, font = defFont, variable = var11, value = 1, command = c1)
    rbutton2 = Radiobutton( frameRB, text = 'TXT file', overrelief = RAISED, font = defFont, variable = var11, value = 2, command = c2)
    rbutton3 = Radiobutton( frameRB, text = ' XLS+TXT', overrelief = RAISED,  font = defFont, variable = var11, value = 3, command = c3)

    frameBt = Frame(window, background = 'white', bd = 2)
    button1 = Button(frameBt, text = 'Запуск',
            # background = "#999",  # фоновый цвет кнопки
            activebackground= "#999",
            foreground = "#0f0",  # цвет текста
            padx = "5",  # отступ от границ до содержимого по горизонтали
            pady = "4",  # отступ от границ до содержимого по вертикали
            font = "10",  # высота шрифта
            overrelief = RAISED,
            command = p1)
    button2 = Button(frameBt, text = 'Отмена',
             background="#777",  # фоновый цвет кнопки
             foreground="#f00",  # цвет текста
             padx="5",  # отступ от границ до содержимого по горизонтали
             pady="4",  # отступ от границ до содержимого по вертикали
             font="10",  # высота шрифта
             overrelief=RAISED,
             command = p2)

    frameLb.pack(side = 'right', fill = 'y', expand = TRUE)
    listbox1.pack(side = 'right', fill = 'y', expand = TRUE)

    frameCB.pack(side = 'left')
    check1.pack(side = 'top')
    check2.pack(side = 'top')
    check3.pack(side = 'top')
    check4.pack(side = 'top')
    check5.pack(side = 'top')
    check6.pack(side = 'top')
    check7.pack(side = 'top')
    check8.pack(side = 'top')
    check9.pack(side = 'top')
    check10.pack(side = 'top')

    frameRB.pack(side = 'top')
    rbutton1.pack(side = 'top', fill = 'both', expand = TRUE)
    rbutton2.pack(side = 'top',fill = 'both', expand = TRUE)
    rbutton3.pack(side = 'top',fill = 'both', expand = TRUE)

    frameBt.pack(side = 'bottom')
    button1.pack(side = 'left')
    button2.pack(side = 'right')

    frameO.pack(side = 'bottom', expand = TRUE)

    frameT.pack(side = 'top', expand = TRUE)
    textT.pack(side = 'top', fill = 'both')

    frameX.pack(side='top', expand=TRUE)
    textX.pack(side = 'top', fill = 'both')

    if var11.get() != 3:
        window.after(100, textX.pack_forget)

    window.wait_window(window)
    return ttt, tttX, vv, ls

def GERAN_UTRAN_Capabilities(Slist, geranPS, geranCS, UTRA, corr, geranoutSw, utranoutSW):
    """
    This for collect founded in the Slist Patt values to Flist
    :param  Slist: input list of Str
            geranPS: search pattern for geranPS caps
            geranCS: search pattern for geranCS caps
            UTRA: search pattern for UTRAN caps
            corr: corr_eNB_GU, Geran/UTRAN capabilities position correction for eNB
    Int.vars: smin,smax, pos, pos2, pos3, geranPS_pos, geranCS_pos, UTRAN_pos:int
    :return: Flist: list of collected values
    e.g. GERAN_UTRAN_Capabilities(s, Patt_geranPS, Patt_geranCS, Patt_UERATcap, corr_eNB_GU, geran_out, utran_out)
    return = tuple  of collected values [GeranPStxt, GeranCStxt, UTRANtxt]
    """
# изначально примем начало блоков GeRAN PS/CS//UTRAN = последней строке трассировки
    UTRAN_pos = geranCS_pos = geranPS_pos = len(Slist)
    logger.info('Пытаемся найти истинные начала блоков GeRAN PS/CS//UTRAN')
    # Init returned values by 'Unknown;
    GeranPStxt = GeranCStxt = UTRANtxt = 'Unknown'

    for i in range(0, len(Slist)):
        pos = Slist[i].find(geranPS)
        if (pos >= 0):
            geranPS_pos = i
            break
    if geranPS_pos == len(Slist):
        logger.warning("geranPS block is absent")

    for i in range(0, len(Slist)):
        pos = Slist[i].find(geranCS)
        if (pos >= 0):
            geranCS_pos = i
            break
    if geranCS_pos == len(Slist):
        logger.warning("geranCS block is absent")

    for i in range(0, len(Slist)):
        pos = Slist[i].find(UTRA)
        if (pos >= 0):
            UTRAN_pos = i
            break
    if UTRAN_pos == len(Slist):
        logger.warning("UTRAN   block is absent")

    # упорядочиваем блоки
    smin = min(geranPS_pos, geranCS_pos, UTRAN_pos)
    smax = max(geranPS_pos, geranCS_pos, UTRAN_pos)
    ##
    logger.info('Ищем строку капабилити в блоке UTRAN')

    if (utranoutSW and (UTRAN_pos < len(Slist))):
        if UTRAN_pos == smax:
            k = len(Slist)
            logger.info("<UTRAN block is 3rd>")
        elif UTRAN_pos in range(smin + 1, smax):
            k = smax
            logger.info("<UTRAN block is 2nd>")
        else:  # UTRAN_pos == smin
            if geranPS_pos > geranCS_pos:
                k = geranCS_pos
                logger.info("<UTRAN block is 1st, geranCS is 2nd>")
            else:
                k = geranPS_pos
                logger.info("<UTRAN block is 1st, geranPS is 2nd>")
        endsearch = min(k+1, len(Slist))
        for i in range(UTRAN_pos, endsearch):
            pos2 = Slist[i].find(Patt_UERATcap)  # позиция начала паттерна в строке
            if (pos2 >= 0):
                pos3 = Slist[i][pos2+17+corr:].find(" --")  # позиция конца паттерна в строке ***************************************************
                # pos3 = Slist[i].find(" (")  # позиция конца паттерна в строке
                if pos3 < 0:
                    UTRANtxt = Slist[i][pos2+17+corr:]
                    utrangeranbinary = False
                    Stail = s[i][-20:]
                    logger.warning('2g/3g capabilities binary output disabled due to UTRAN capabilities unqualified, tail is %s', Stail)
                else:
                    UTRANtxt = Slist[i][pos2+17+corr: pos2+17+pos3+corr]
                break
     ##
    logger.info('Поиск капабилити для GeranPS')
    if geranoutSw and (geranPS_pos < len(Slist)):
        if geranPS_pos == smax:
            k = len(Slist)
            logger.info("<geranPS block is 3rd>")
        elif geranPS_pos in range(smin + 1, smax):
            k = smax
            logger.info("<geranPS block is 2nd>")
        else:  # geranPS_pos == smin
            if geranCS_pos > UTRAN_pos:
                k = UTRAN_pos
                logger.info("<geranPS block is 1st, UTRAN is 2nd>")
            else:
                k = geranCS_pos
                logger.info("<geranPS block is 1st, geranCS is 2nd>")

        ##
        logger.info('Ищем строку капабилити в блоке geranPS')
        endsearch = min(k + 1, len(Slist))
        for i in range(geranPS_pos, endsearch):
            pos2 = Slist[i].find(Patt_UERATcap)  # позиция начала паттерна в строке
            if (pos2 >= 0):
                pos3 = Slist[i][pos2+17+corr:].find(" ---- ")  # позиция конца паттерна в строке *************************************************************
                if pos3 < 0:
                    GeranPStxt = Slist[i][pos2+17+corr:]
                    utrangeranbinary = False
                    Stail = s[i][-20:]
                    logger.warning('2g/3g capabilities binary output disabled due to GeranPS capabilities unqualified, tail is %s', Stail)
                else:
                    GeranPStxt = Slist[i][pos2+17+corr:pos2+17+pos3+corr]
                break
    ##
    logger.info('Поиск капабилити для GeranCS')
    if geran_out and geranCS_pos < len(s):
        if geranCS_pos == smax:
            k = len(s)
            logger.info("<geranCS block is 3rd>")
        elif geranCS_pos in range(smin + 1, smax):
            k = smax
            logger.info("<geranCS block is 2nd>")
        else:  # geranCS_pos == smin
            if geranPS_pos > UTRAN_pos:
                k = UTRAN_pos
                logger.info("<geranCS block is 1st, UTRAN is 2nd>")
            else:
                k = geranPS_pos
                logger.info("<geranCS block is 1st, geranPS is 2nd>")
            ##
        logger.info('Ищем строку капабилити в блоке geranCS')
        endsearch = min(k + 1, len(Slist))
        for i in range(geranCS_pos, endsearch):
            pos2 = Slist[i].find(Patt_UERATcap)  # позиция начала паттерна в строке
            if (pos2 >= 0):
                pos3 = Slist[i][pos2+17+corr:].find(" ---- ")  # позиция конца паттерна в строке *************************************************************
                if pos3 < 0:
                    GeranCStxt = Slist[i][pos2+17+corr:]
                    utrangeranbinary = False
                    Stail = s[i][-20:]
                    logger.warning('2g/3g capabilities binary output disabled due to GeranCS capabilities unqualified, tail is %s', Stail)
                else:
                    GeranCStxt = Slist[i][pos2+17+corr:pos2+17+pos3+corr]
                break
    return [GeranPStxt, GeranCStxt, UTRANtxt]

#----------------------------------------------------------------------------------------------------------------------
#---------------------------------- M A I N ---------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------------------
# fn -  полное имя входного файла любого типа
# fnO - полное имя выходного текстового файла
# fnX - полное имя выходного Excel файла
# f0  -  файл ввода-вывода

FD_win = tkinter.Tk()
FD_win.title("Msg2Cap")
# FD_win.wm_withdraw() # this completely hides the FD_win window
# FD_win.iconify() # this will move the FD_win window to a minimized icon.

t1 = PrettyTable()
t2 = PrettyTable()
if len(sys.argv) < 2:  # нет параметров
    # Call dialogue box to ask input file with capability_info trace
    fn = tkinter.filedialog.askopenfilename()
    FD_win.destroy()
else:
    fn = sys.argv[1]
cwd = os.getcwd()  # текущий каталог
if os.path.exists(fn):  # если файл трассировки существует
    f = open(fn, 'r')
    Fname, ext = os.path.splitext(fn)
    # считываем текстовый файл в список
    # Should be '.txt'
    if ('txt' in ext):
        # Add prefix 'Parsed_' to the input file name for the ouput (parsed) filename (with the same file path)
        # fnO = os.path.join(os.path.split(fn)[0], 'Parsed_' + os.path.split(fn)[1])
        fnO = os.path.splitext(fn)[0] + '_parsed.txt'
        logging.info('Output text file: %s', fnO)

        fn1, fn1X, fl, sl = GUI(fnO, os.path.splitext(fn)[0] + '_parsed.xlsx')  #ttt, tttX, vv, ls
        #switch to the parser's loger
        logger = logging.getLogger('_Parser')

# ----------------------------------------------------------------------------------------------------------------------
        # Translate  GUI outputs
        fgi_out = bool(fl[0])
        geran_out = bool(fl[1])
        utran_out = bool(fl[2])
        utrangeranbinary = bool(fl[3])
        R14_enabled = bool(fl[4])
        table_format = bool(fl[5])
        BandsFilter = bool(fl[6])

        if fl[10] == 1: # Screen output only
            txt_output = False
            Excel_output = False

        elif fl[10] == 2: # txt output only
            txt_output = True
            fn1 = fn1.rstrip('\n')
            if fn1 == '':
                fnO = os.path.splitext(fn)[0] + '_parsed.txt'
                logging.warning('Returned NO output TXT file name, it changed to: %s', fnO)
            else:
                fnO = fn1
                logging.info('Output TXT file changed in GUI to: %s', fn1)
            Excel_output = False

        elif fl[10] == 3: #XLS output + TXT output
            txt_output = True
            Excel_output = True
            fn1X = fn1X.rstrip('\n')
            if fn1X == '':
                fnX = os.path.splitext(fn)[0] + '_parsed.xlsx'
                logging.warning('Returned NO output XLS file name, it changed to: %s', fnX)
            else:
                fnX = fn1X
                logging.info('Output XLS file changed in GUI to: %s', fnX)
        else: # Unknown value returned from GUI
            logging.error('Output Rbutton = %s is out of range', fl[10])
            assert(1<=fl[10]<=3)
        logging.info('Output text file changed to: %s', fn1)
        logging.info('Selected entries: %s', sl)
        logging.info('Switches: %s', fl)
# ----------------------------------------------------------------------------------------------------------------
        s = f.readlines()
        f.close()
    # Should be '.xls','.xlsx','.xlsm'
    elif ('xls' in ext):
        # Excel_output = True
        XLwithmacro = ('.xlsm','.xltm')
        XLwithoutmacro = ('.xls','.xlsx','.xltx')
        # меняем логгер на XLS
        logger = logging.getLogger('XLSproc')
        # Входная трассировка в файле Excel на вкладке Capabilities
        logger.info('Open workbook: %s', fn)
        if (ext in XLwithmacro):
            book = openpyxl.load_workbook(fn, keep_vba=True)
        elif (ext in XLwithoutmacro):
            book = openpyxl.load_workbook(fn)
        else:
            logging.error('Extension %s is not known', ext)
            exit(1)

        fnO = os.path.splitext(fn)[0]+ '_parsed.txt'
        fnX = os.path.splitext(fn)[0]+'_parsed'+ext
        logging.info('Output XLS file: %s', fnX)


        fn1, fn1X, fl, sl = GUI(fnO,fnX)
        #switch to the parser's loger
        logger = logging.getLogger('_XLSproc')

        # Translate  GUI outputs
        # ----------------------------------------------------------------------------------------------------------------------
        fgi_out = bool(fl[0])
        geran_out = bool(fl[1])
        utran_out = bool(fl[2])
        utrangeranbinary = bool(fl[3])
        R14_enabled = bool(fl[4])
        table_format = bool(fl[5])
        BandsFilter = bool(fl[6])

        if fl[10] == 1:  # Screen output only
            txt_output = False
            Excel_output = False

        elif fl[10] == 2:  # txt output only
            txt_output = True

            fn1 = fn1.rstrip('\n')
            if fn1 == '':
                fnO = os.path.splitext(fn)[0] + '_parsed.txt'
                logging.warning('Returned NO output TXT file name, it changed to: %s', fnO)
            else:
                fnO = fn1
                logging.info('Output TXT file changed in GUI to: %s', fn1)
            Excel_output = False

        elif fl[10] == 3:  # XLS output + TXT output
            txt_output = True
            Excel_output = True

            fn1X = fn1X.rstrip('\n')
            if fn1X == '':
                fnX = os.path.splitext(fn)[0] + '_parsed.xlsx'
                logging.warning('Returned NO output XLS file name, it changed to: %s', fnX)
            else:
                fnX = fn1X
                logging.info('Output XLS file changed in GUI to: %s', fnX)

        else:  # Unknown value returned from GUI
            logging.error('Output Rbutton = %s is out of range', fl[10])
            assert (1 <= fl[10] <= 3)
        logging.info('Output text file changed to: %s', fnO)
        logging.info('Output XLS file: %s', fnX)
        logging.info('Selected entries: %s', sl)
        logging.info('Switches: %s', fl)
        # ----------------------------------------------------------------------------------------------------------------

        logger.info('Loading device capabilities from file %s', fn)
        logger.info('Sheets: %s', book.sheetnames)
        sheetIn = "Capabilities"
        sheetOut = "Parsed_Capabilities"
        if sheetIn in book.sheetnames:
            # Меняем вкладку
            sheet = book[sheetIn]
        else:
            logger.warning('Вкладка %s в таблице не найдена', sheetIn)
            exit(0)
        logger.info('Вкладка %s', sheet.title)
        s = []
        row: object
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value != None:
                    s.append(str([cell.value ]).rstrip("'/]"))
        # FD_win.destroy()
    else:
        logger.error('File type %s undefined', ext)
        exit(0)
else:  # Файл трассировки не найден
    logger.error("File %s not found %s >", os.path.splitext(fn)[0], cwd)
    exit

# В дальнейшем вывод производится в заданный файл вывода (стандартный вывод на экран или в текстовый файл на диске)+++ или в файл Excel
if txt_output:  # <===================================================================================================================
    ff = fnO
else:
    ff = 'null'
    f0 = sys.stdout
with open(ff, 'w') as fO:
    if not(txt_output):
        fO = sys.stdout
    logger.info("Трассировка %s содержит %s строк", fn, len(s))
    print("Трассировка", fn, " содержит ", len(s), " строк", file=fO)
    print("", file=fO)
    ##
    logger.info('Определим, трассировка декодирована TraceViewer-ом из tmf или самой eNB')
    cnt_TMF = 0  # Обнуляем счетчик нахождений паттерна TMF
    cnt_eNB = 0  # Обнуляем счетчик нахождений паттерна eNB
    # Перебираем все строки для сбора бэндов EUTRA
    for i in range(0, len(s)):
        pos = s[i].find(Patt_SBtmf)
        if (pos >= 0):
            cnt_TMF += 1
        else:
            pos = s[i].find(Patt_SBeNB)
            if (pos >= 0):
                cnt_eNB += 1
    corr_eNB_FGI = 0  # Коррекция позиции  FGI для tmf не требуется
    corr_eNB_GU = 0  # Коррекция позиций ueCapabilityRAT для tmf не требуется
    if cnt_TMF > cnt_eNB:
        logger.info("TMF format decided [%s : %s]", cnt_TMF, cnt_eNB)
        logger.info("TMF format decided")
        Patt_SB = Patt_SBtmf
        Patt_UtraBands = Patt_UtraBands_tmf
        Patt_GeRANBands = Patt_GeRANBands_tmf
        Patt_FGI8 = Patt_FGI8_tmf
        Patt_FGI9 = Patt_FGI9_tmf
        Patt_FGI10 = Patt_FGI10_tmf
        Patt_geranCS = Patt_geranCS_tmf
        Patt_geranPS = Patt_geranPS_tmf
        Patt_UTRA = Patt_UTRA_tmf
        Patt_UERATcap = Patt_UERATcap_tmf
        Patt_BWclassUL = Patt_BWclassUL_tmf
        Patt_BWclassDL = Patt_BWclassDL_tmf
        Patt_MIMO = Patt_MIMO_tmf
        Patt_CC10 = Patt_CC10_tmf
        Patt_Band10 = Patt_Band10_tmf
        TMF = True
        eNB = False
    else:
        logger.info("eNB format decided")
        logger.info("eNB format decided [%s : %s]", cnt_TMF, cnt_eNB)
        Patt_SB = Patt_SBeNB
        Patt_UtraBands = Patt_UtraBands_eNB
        Patt_GeRANBands = Patt_GeRANBands_eNB
        Patt_FGI8 = Patt_FGI8_eNB
        Patt_FGI9 = Patt_FGI9_eNB
        Patt_FGI10 = Patt_FGI10_eNB
        Patt_geranCS = Patt_geranCS_eNB
        Patt_geranPS = Patt_geranPS_eNB
        Patt_UTRA = Patt_UTRA_eNB
        Patt_UERATcap = Patt_UERATcap_eNB
        corr_eNB_FGI = 1  # Коррекция позиции  FGI = на 1 позицию
        corr_eNB_GU = 2  # на 2 позиции правее
        Patt_BWclassUL = Patt_BWclassUL_eNB
        Patt_BWclassDL = Patt_BWclassDL_eNB
        Patt_MIMO = Patt_MIMO_eNB
        Patt_CC10 = Patt_CC10_eNB
        Patt_Band10 = Patt_Band10_eNB
        Patt_StartComb = Patt_StartCombeNB
        Patt_endComb = Patt_endCombeNB
        Patt_UEacc = Patt_UEacc_eNB
        Patt_UEcat = Patt_UEcat_eNB
        eNB = True
        TMF = False
    print("\n", file=fO)
    ##
    logger.info('Active SWITCHes:')
    logger.info('FGI_out = %s',fgi_out)
    logger.info('geran_out = %s',geran_out)
    logger.info('utran_out = %s',utran_out)
    logger.info('utrangeranbinary = %s',utrangeranbinary)
    logger.info('R14_enabled = %s',R14_enabled)
    logger.info('table_format = %s',table_format)
    logger.info('BandsFilter = %s',BandsFilter)
    logger.info('txt_output = %s',txt_output)
    logger.info('Excel_output = %s',Excel_output)
    ##
    logger.info('Перебираем все строки для поиска UE Access Stratum')
    UEaccS = CollectPatt(s, Patt_UEacc,0)
    print("Найденные 3gpp Access Stratum:", len(UEaccS), file=fO)
    print("=================================", file=fO)
    for i in range(0, len(UEaccS)):
        print(UEaccS[i], end='', file=fO)
    print("\n", file=fO)
    ##
    logger.info('Перебираем все строки для сбора категорий')
    UEcats = CollectPatt(s, Patt_UEcat,2)
    print("Найденные категории LTE:", len(UEcats), file=fO)
    print("=================================", file=fO)
    S_UEcat = []
    S_UEcatUL = []
    S_UEcatDL = []
    for i in range(0, len(UEcats)):
        print(UEcats[i], end='\n', file=fO)
        if 'DL' in UEcats[i]:
            S_UEcatDL.append(int(valfrombrackets(UEcats[i])))
        elif 'UL' in UEcats[i]:
            S_UEcatUL.append(int(valfrombrackets(UEcats[i])))
        else:
            S_UEcat.append(int(valfrombrackets(UEcats[i])))
    print("\n", file=fO)
    S_UEcat.sort()
    S_UEcatDL.sort()
    S_UEcatUL.sort()
    logger.info('LTE Categories: %s', S_UEcat)
    logger.info('LTE DL Categories: %s', S_UEcatDL)
    logger.info('LTE UL Categories: %s', S_UEcatUL)
    ##
    S_EUTRA = [] #  номера LTE bands, Обнуляем начальное значение несущих EUTRA
    nb = 0
    logger.info('Перебираем все строки для сбора бэндов EUTRA')
    for i in range(0, len(s)):
        pos = s[i].find(Patt_SB)
        if (pos >= 0):
            Bands.append([0, 0, 0])
            eUtraBands.append("Band  " + (valfrombrackets(s[i][pos + 15:])))
            S_EUTRA.append(int (valfrombrackets (s[i][pos + 15:])))
            Bands[nb][0] = int(valfrombrackets(s[i][pos + 15:]))
            nb += 1
    S_EUTRA.sort()
    logger.info('LTE bands: %s', S_EUTRA)

    NumSBr12 = -1  # инициируем переменную количества бэндов с расширением для R12 значением -1 чтобы она могла быть индексом бэнда
    SBr12Found = False  # инициируем переменную наличия блока расширения для R12 значением False
    # для начала ипринимаем за паттерны UL/DL паттерны для блока V12
    Patt_dl = Patt_dl256
    Patt_ul = Patt_ul64
    for i in range(0, len(s) - 2):
        pos = s[i].find(Patt_SBr12)
        if (pos >= 0):
            SBr12Found = True  # блок модуляций для R12xx присутствует
            NumSBr12 += 1  # количество блоков модуляций для r12 ++
            #print("---SupportedBandEUTRA-v12 =", NumSBr12, file=fO)
            # Если найденный блок v1250, переопределим паттерны для переиспользования кода
            pos = s[i].find(Patt_SBr1250)
            if (pos >= 0):
                Patt_dl = Patt_dl256_1250
                Patt_ul = Patt_ul64_1250
            if TMF:
                Patt_dl = Patt_dl256tmf
                Patt_ul = Patt_ul64tmf
                #   Поиск DL256
            # if ((s[i + 1].find(Patt_dl) >= 0) or (s[i + 2].find(Patt_dl) >= 0)):
            if ((Patt_dl in s[i + 1]) or (Patt_dl in s[i + 2])):
                 eUtraBands[NumSBr12] = eUtraBands[NumSBr12] + " 256QAM DL"
                 Bands[NumSBr12][1] = 8  # bits
#                if (Bands[NumSBr12][0] = PrimaryEUTRABand)
                 if (Bands[NumSBr12][0] in (PrimaryEUTRABand, SecondaryEUTRABands)):
                    S256DL = True #256QAM for DL of some used bands (primary or secondary)
            else:
                eUtraBands[NumSBr12] = eUtraBands[NumSBr12] + " 64QAM  DL"
                Bands[NumSBr12][1] = 6  # bits
            #   Поиск UL64
            if ((s[i + 1].find(Patt_ul) >= 0) or (s[i + 2].find(Patt_ul) >= 0)):
                eUtraBands[NumSBr12] = eUtraBands[NumSBr12] + " 64QAM UL"
                Bands[NumSBr12][2] = 6  # bits
            else:
                eUtraBands[NumSBr12] + eUtraBands[NumSBr12] + " 16QAM UL"
                Bands[NumSBr12][2] = 4  # bits
    print("\n", file=fO)

    if not (SBr12Found):  # блок модуляций для r12 отсутствует
        for j in range(nb):
            eUtraBands[j] = eUtraBands[j] + " 64QAM DL 16QAM UL \n"
            Bands[j][1] = 6  # bits
            Bands[j][2] = 4  # bits

    print("Найденные EUTRA bands ( с расширением до v12.50):", len(eUtraBands), file=fO)
    print("====================================================", file=fO)
    for i in range(0, len(eUtraBands)):
        print(eUtraBands[i], end='\n', file=fO)
    print("\n", file=fO)
    ##
    S_UTRA = []
    logger.info('Перебираем все строки для сбора бэндов UTRA')
    S_UTRAN = CollectPatt(s, Patt_UtraBands, 28)
    S_UTRAN.sort()
    print("Найденные UTRA bands:", len(S_UTRAN), file=fO)
    print("========================", file = fO)
    for i in range(0, len(S_UTRAN)):
        S_UTRA.append(int(valfrombrackets(S_UTRAN[i]))+1)
        print(S_UTRA[i],'\t',S_UTRAN[i], end='\n', file=fO)
    print("\n", file=fO)
    logger.info('UTRA bands: %s', S_UTRA)

    # Перебираем все строки для сбора бэндов GERAN
    S_GERAN = []
    for i in range(0, len(s)):
        pos = s[i].find(Patt_GeRANBands)
        if (pos >= 0):
            GeRANBands.append(valfromGSM(s[i]))
            S_GERAN.append(valfromGSM(s[i]))
    print("Найденные GeRAN bands:", len(GeRANBands), file=fO)
    print("========================", file = fO)
    for i in range(0, len(GeRANBands)):
        print(GeRANBands[i], end='\n', file=fO)
    print("\n", file=fO)
    S_GERAN.sort()
    logger.info('GERAN bands: %s', S_GERAN)

    if fgi_out:
        # Поиск FGI r8
        logger.info('Поиск FGI r8')
        for i in range(0, len(s)):
            pos = s[i].find(Patt_FGI8)
            if (pos >= 0):
                FGI8Txt = valFGIfromquotes(s[i])
                if len(FGI8Txt) == 32:
                    print("         0               1               ", file=fO)
                    print("         0123456789ABCDEF0123456789ABCDEF", file=fO)
                    print("         +---------------+---------------+", file=fO)
                    print("FGI r8: ", FGI8Txt, file=fO)
                else:
                    print("No FGI r8 found", file=fO)
                    logger.warning("No FGI r8 found")
        print("\n", file=fO)
        ##
        logger.info('Поиск FGI r9')
        for i in range(0, len(s)):
            pos = s[i].find(Patt_FGI9)
            if (pos >= 0):
                FGI9Txt = valFGIfromquotes(s[i])
                if len(FGI9Txt) == 32:
                    print("         0               1               ", file=fO)
                    print("         0123456789ABCDEF0123456789ABCDEF", file=fO)
                    print("         +---------------+---------------+", file=fO)
                    print("FGI r9: ", FGI9Txt, file=fO)
                else:
                    print("No FGI r9 found", file=fO)
                    logger.warning("No FGI r9 found")
        print("\n", file=fO)
        ##
        logger.info('Поиск FGI r10')
        for i in range(0, len(s)):
            pos = s[i].find(Patt_FGI10)
            if (pos >= 0):
                FGI10Txt = valFGIfromquotes(s[i])
                if len(FGI10Txt) == 32:
                    print("         0               1               ", file=fO)
                    print("         0123456789ABCDEF0123456789ABCDEF", file=fO)
                    print("         +---------------+---------------+", file=fO)
                    print("FGI r10:", FGI10Txt, file=fO)
                else:
                    print("No FGI r10 found", file=fO)
                    logger.warning("No FGI r10 found")
        print("\n", file=fO)
    ##
    logger.info('Поиск капабилити 2g/3g')
    [GeranPScapTxt,GeranCScapTxt,UTRANcapTxt]=GERAN_UTRAN_Capabilities(s, Patt_geranPS, Patt_geranCS, Patt_UTRA, corr_eNB_GU, geran_out, utran_out)
    if (len(UTRANcapTxt) > 1) and utran_out:
        print("UTRAN capabilities: \n =0x", UTRANcapTxt, sep='', file=fO)
        if utrangeranbinary:
            UTRANcapBits = Conv2Bits(UTRANcapTxt)
            print("\nBinary = ", end='', file=fO)
            for i in range(len(UTRANcapBits)):
                print(UTRANcapBits[i], end='',file=fO)
            print("\n", file=fO)
    print("\n", file=fO)
    if (len(GeranPScapTxt) > 1) and geran_out:
        print("\nGERAN PS capabilities: \n =0x", GeranPScapTxt, sep='', file=fO)
        if utrangeranbinary:
            GeranPScapBits = Conv2Bits(GeranPScapTxt)
            print("\nBinary = ", end='', file=fO)
            for i in range(len(GeranPScapBits)):
                print(GeranPScapBits[i], end='', file=fO)
            print("\n", file=fO)
    print("\n", file=fO)
    if (len(GeranCScapTxt) > 1) and geran_out:
        print("GERAN CS capabilities: \n =0x", GeranCScapTxt, sep='', file=fO)
        if utrangeranbinary:
            GeranCScapBits = Conv2Bits(GeranCScapTxt)
            print("\nBinary = ", end='', file=fO)
            for i in range(len(GeranCScapBits)):
                print(GeranCScapBits[i], end='', file=fO)
            print("\n", file=fO)
    print("\n", file=fO)

    # Костыль для поиска UL 256QAM для R14
    CCr14Mod = []  # Массив/список для хранения расширений UL 256QAM для R14
    NumCCr14 = 0  # инициируем переменную количества комбинаций с UL 256QAM для R14
    CCr14Found = False  # инициируем переменную наличия блока расширения для R14 значением False

    for i in range(0, len(s)):
        pos = s[i].find(Patt_ul256r14)
        if (pos >= 0):
            CCr14Found = True  # блок модуляций для r14 присутствует
            NumCCr14 += 1  # количество блоков модуляций для r14 ++
            #print("---SupportedCCEUTRA-r14 =", NumCCr14, file=fO)
            if 'supported' in s[i]:
                CCr14Mod.append(8)  # bits
            else:
                CCr14Mod.append(None)
    print("\n", file=fO)
    if CCr14Found:
        logger.info('Найдена поддержка компонент с UL256QAM r14.3')
    logger.info('Поиск начала блока комбинаций несущих r10')
    i = 0
    while (i < len(s) and s[i].find(Patt_StartComb) < 0):
        i += 1
    Comb_Start = i
    logger.info('Поиск конца блока комбинаций несущих r10')
    while (i < len(s) and s[i].find(Patt_endComb) < 0):
        i += 1
    Comb_End = i
    if (Comb_End > Comb_Start):
        logger.info('Перебираем ВСЕ строки блока комбинаций несущих')
        NCCs = 0
        Ncarr = 0
        for i in range(Comb_Start, Comb_End + 1):
            if (s[i].find(Patt_bc) >= 0):
                NCCs += 1
            endcc = True
            while (endcc):
                if (s[i].find(Patt_CC10) >= 0):
                    CCs.append([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])  # Добавляем пустую запись для первой несущей в найденной комбинации несущих
                    Ncarr += 1  # 0,1,2,3,4,5,6,7,8,9,A - поля
                    # cбор и заполнение параметров для найденной несущей
                    CCs[Ncarr - 1][0] = NCCs
                    CCs[Ncarr - 1][1] = int(valfrombrackets(s[i][pos + 20:]))
                    # находим параметры модуляций UL/DL для текущего Band  в  списке Bands  и копируем их в поля 1..3
                    for j in range(0, len(eUtraBands)):
                        if (Bands[j][0] == CCs[Ncarr - 1][1]):
                            CCs[Ncarr - 1][2] = Bands[j][1]
                            CCs[Ncarr - 1][3] = Bands[j][2]
                            # contigous carriers in UL -->[5]
                if s[i].find(Patt_BWclassUL) >= 0:
                    temp_int = int(valfrombrackets(s[i][pos + 30:]))
                    CCs[Ncarr - 1][5] = temp_int
                    if (temp_int == 1):  # class "B"
                        CCs[Ncarr - 1][5] = 2
                        CCs[Ncarr - 1][8] = 10
                    elif (temp_int == 0):  # class "A"
                        CCs[Ncarr - 1][5] = 1
                        CCs[Ncarr - 1][8] = 20
                    else:  # classes {C,D...Z}
                        CCs[Ncarr - 1][5] = temp_int
                        CCs[Ncarr - 1][8] = 20
                    # contigous carriers in DL --> [4]
                if s[i].find(Patt_BWclassDL) >= 0:
                    temp_int = int(valfrombrackets(s[i][pos + 30:]))
                    if (temp_int == 1):  # class "B"
                        CCs[Ncarr - 1][4] = temp_int + 1
                        CCs[Ncarr - 1][7] = 10
                    elif (temp_int == 0):  # class "A"
                        CCs[Ncarr - 1][4] = temp_int + 1
                        CCs[Ncarr - 1][7] = 20
                    else:  # classes {C,D...Z}
                        CCs[Ncarr - 1][4] = temp_int
                        CCs[Ncarr - 1][7] = 20
                if s[i].find(Patt_MIMO) >= 0:
                    temp_int = int(valfrombrackets(s[i][pos + 45:]))
                    if (temp_int == 0):
                        CCs[Ncarr - 1][6] = 2
                    elif (temp_int == 1):
                        CCs[Ncarr - 1][6] = 4 # 4 layers
                        if (CCs[Ncarr-1][1] in (PrimaryEUTRABand,SecondaryEUTRABands)):
                            S_4x4 = True
                    elif (temp_int == 2):
                        CCs[Ncarr - 1][6] = 8 # 8 layers
                        if (CCs[Ncarr - 1][2] in (PrimaryEUTRABand, SecondaryEUTRABands)):
                            S_8x8 = True  # 8 layers for some of use band(s)
                    if (CCs[Ncarr - 1][1] > 32) & (CCs[Ncarr - 1][1] < 49):  # TDD ULDL_config#3 (DL/UL=6/3): DL = 0.6*0.625=0.375 | UL = 0.3*0.625 = 0.1875
                        IsFDD = False
                        CCs[Ncarr - 1][9] = int(
                            0.375 * CCs[Ncarr - 1][2] * CCs[Ncarr - 1][4] * CCs[Ncarr - 1][6] * CCs[Ncarr - 1][7])
                        CCs[Ncarr - 1][10] = int(0.1875 * CCs[Ncarr - 1][3] * CCs[Ncarr - 1][5] * CCs[Ncarr - 1][8])
                    else:  # FDD
                        IsFDD = True
                        CCs[Ncarr - 1][9] = int(
                            0.625 * CCs[Ncarr - 1][2] * CCs[Ncarr - 1][4] * CCs[Ncarr - 1][6] * CCs[Ncarr - 1][7])
                        CCs[Ncarr - 1][10] = int(0.625 * CCs[Ncarr - 1][3] * CCs[Ncarr - 1][5] * CCs[Ncarr - 1][8])
                endcc = False
        #       print("\n",file = fO)
        print("Найдено ", NCCs, " комбинаций несущих:", file=fO)
        #       print("\n",file = fO)
        if NCCs > 0:
            nccUL_r14=0 # обнуляем номер UL компоненты для добавки данных UL модуляций R1430
            #Вносим коррекции в UL при наличии блока UL 256QAM R14 и включенного свитча обработки R14
            if (R14_enabled and CCr14Found):
                logger.info('Обрабатываем расширение модуляций для R14')
                for l in range(Ncarr):
                    # если у текущей компоненты UL активен (имеет не менее 1 пространственного канала)
                    # и для этой комбинации разрешен 256QAM UL (не None)
                    if (CCs[l][5]>0 and CCr14Mod[nccUL_r14]>0):
                        CCs[l][3] = CCr14Mod[nccUL_r14]
                        CCs[l][10] = int(0.625 * CCs[l][3] * CCs[l][5] * CCs[l][8])
                        nccUL_r14 += 1
                        if CCs[l][2] in (PrimaryEUTRABand, SecondaryEUTRABands):
                            S256UL = True
            if CCr14Found:
                logger.info('Найдено %s компонент с UL 256QAM',nccUL_r14)
            ##
            logger.info('Создаем список комбинаций ')
            for i in range(NCCs):  # комбинации
                # заполняем список комбинаций суммами компонентных несущих
                CCsCC.append([i + 1, 0, 0])
                for j in range(Ncarr):  # компоненты
                    if (CCs[j][0] == i + 1):
                        CCsCC[i][1] = CCsCC[i][1] + CCs[j][9]
                        CCsCC[i][2] = CCsCC[i][2] + CCs[j][10]
            if table_format:
                t1.field_names = ["№комб", "DL Mbps", "UL_Mbps"]
                #       Alignments
                t1.align["№комб"] = "r"
                t1.align["DL Mbps"] = "r"
                t1.align["UL_Mbps"] = "r"
            else:
                print("№комб.", " ", "DL Mbps", "      ", "UL_Mbps", file=fO)
            if not (table_format):
                print("================================", file=fO)
            ##
            logger.info('Вывод всех комбинаций')
            for i in range(NCCs):
                if not (table_format):
                    print(CCsCC[i][0], "\t", CCsCC[i][1], "\t\t", CCsCC[i][2], end="\n", file=fO)
                else:
                    t1.add_row([CCsCC[i][0], CCsCC[i][1], CCsCC[i][2]])
            if table_format:
                print(t1, file=fO)
        print("\n", file=fO)
        ##
        logger.info('Вывод списка компонентных несущих')
        if Ncarr > 0:
            print("Найдено ", Ncarr, " компонентов несущих:", file=fO)
            print("Для TDD расчет ожидаемой скорости выполняется для конфигурации TDD #3 (6DL/3UL таймслотов)", file=fO)
            #            print("\n",file = fO)
            if not (table_format):
                print("СС#     №комб. Band#   DL_бит  UL_бит  №DL     №UL    MIMO    DLbwMHz  ULbwMHz DL,Mbps UL.Mbps",
                      file=fO)
                print(
                    "=================================================================================================",
                    file=fO)
            else:
                t2.field_names = ["№комб", "СС#", "Band#", "DL_bits", "UL_bits", "№DL", "№UL", "MIMO", "DLbwMHz",
                                  "ULbwMHz", "DL,Mbps", "UL.Mbps"]
            #           t2.align = ["r","r","r","r","r","r","r","r","r","r","r","r"]
            prevComb = 1
            for i in range(Ncarr):  # компоненты 0...Ncarr-1
                # Ищем S_CA7C
                if ((CCs[i][1] == 7) and (CCs[i][4] == 2)):
                    S_CA7C = True
                if CCs[i][0] > prevComb:  # если текущая комбинация увеличилась
                    temp_int = CCs[i][0] - 1  # комбинация = прошлая, выведем для нее суммарные скорости:
                    if table_format:
                        t2.add_row([temp_int, "SUM", "--", "-", "-", "-", "-", "-", "--", "--", CCsCC[temp_int - 1][1],
                                    CCsCC[temp_int - 1][2]])
                        t2.add_row(["", "", "", "", "", "", "", "", "", "", "", ""])
                    else:
                        if temp_int < 10:  # Alignment (not Table)
                            print("Комб.==", temp_int, "=" * 69, CCsCC[temp_int - 1][1], "===", CCsCC[temp_int - 1][2],
                                  file=fO)
                        else:
                            print("Комб.==", temp_int, "=" * 68, CCsCC[temp_int - 1][1], "===", CCsCC[temp_int - 1][2],
                                  file=fO)
                        print("\n", file=fO)
                prevComb = CCs[i][0]  # сохраняем номер текущей комбинации как предыдущий
                if not (table_format):
                    print(i + 1, "\t", end='', file=fO)  # печать индекса = номер CC компоненты
                    for j in range(11):  # колонки 0..10
                        print(CCs[i][j], "\t", end='', file=fO)  # печать колонок 0..10
                    print("\n", file=fO)
                else:
                    t2.add_row(
                        [CCs[i][0], i + 1, CCs[i][1], CCs[i][2], CCs[i][3], CCs[i][4], CCs[i][5], CCs[i][6], CCs[i][7],
                         CCs[i][8], CCs[i][9], CCs[i][10]])
            temp_int = NCCs - 1  # комбинация = последняя, выведем для нее суммарные скорости:
            if not (table_format):
                if temp_int + 1 < 10:  # Alignment (not Table)
                    print("Комб.==", temp_int + 1, "=" * 69, CCsCC[temp_int][1], "===", CCsCC[temp_int][2], file=fO)
                else:
                    print("Комб.==", temp_int + 1, "=" * 68, CCsCC[temp_int][1], "===", CCsCC[temp_int][2], file=fO)
            else:
                t2.add_row([temp_int + 1, "SUM", "--", "-", "-", "-", "-", "-", "--", "--", CCsCC[temp_int][1],
                            CCsCC[temp_int][2]])
            if table_format:
                print(t2, file=fO)
            print("\n", file=fO)
            logger.info("UE Capabilities parsing finished")
            if Excel_output:
                logger = logging.getLogger('XLSproc')
                if 'txt' in ext:
                    # txt trace should be parsed to the xls output
                    # new  xls file should be created with fnX name+path
                    book = openpyxl.Workbook()
                if "Parsed_Capabilities" not in book.sheetnames:
                    # Add new sheet, fill it and save book copy with new sheet
                    ws1 = book.create_sheet("Parsed_Capabilities")  # insert sheet at the end (by default)
                    logging.info('Вкладка Cap.Info создана')
                else:
                    logging.info('Вкладка Cap.Info уже существует')
                sheet = book["Parsed_Capabilities"]
                sheet.sheet_properties.tabColor = "1072BA"

                sheet['A2'] = 'LTE category DL+UL'
                sheet['B2'] = str(S_UEcat)
                sheet['A3'] = 'LTE category DL'
                sheet['B3'] = str(S_UEcatDL)
                sheet['A4'] = 'LTE category UL'
                sheet['B4'] = str(S_UEcatUL)
                sheet['A5'] = 'EUTRA_Bands'
                sheet['B5'] = str(S_EUTRA)
                sheet['A6'] = 'UTRA_Bands'
                sheet['B6'] = str(S_UTRA)
                sheet['A7'] = 'GeRAN_Bands'
                sheet['B7'] = str(S_GERAN)
                sheet['A8'] = '256QAM for used bands'
                if S256DL and S256UL:
                    sheet['B8'] = 'DL + UL'
                else:
                    if S256DL:
                        sheet['B8'] = 'DL'
                    elif S256UL:
                        sheet['B8'] = 'UL'
                    else:
                        sheet['B8'] = ''
                sheet['B8'].font = Font(name=FSimple,size=12, underline='none', color=DARKBLUE, bold=True, italic=False)

                sheet['A9'] = 'MIMO 4x4 for used bands'
                if S_4x4:
                    sheet['B9'] = 'Supported'
                else:
                    sheet['B9'] = 'Not Supported'
                sheet['B9'].font = Font(name=FBold,size=12, underline='none', color=DARKBLUE, bold=True, italic=False)

                sheet['A10'] = 'CA 7c, 256 QAM, MIMO 4x4'
                S_Qualcomm = (((S256DL or S256UL) and S_4x4) and S_CA7C)
                if S_Qualcomm:
                    sheet['B10'] = 'Supported'
                else:
                    sheet['B10'] = 'Not Supported'
                sheet['B10'].font = Font(name=FBold,size=12, underline='none', color=DARKBLUE, bold=True, italic=False)
                RowInSheet = 15
                sheet.cell(row=RowInSheet, column=1, value='UE access stratum:')
                sheet.cell(row=RowInSheet, column=2, value= str(UEaccS))
                RowInSheet=16
                sheet.cell(row=RowInSheet, column=1, value='UE categories:')
                sheet.cell(row=RowInSheet, column=2, value= str(UEcats))
                RowInSheet=18
                sheet.cell(row=RowInSheet, column=2, value= str(FGILine1))
                RowInSheet=19
                sheet.cell(row=RowInSheet, column=2, value= str(FGILine2))
                RowInSheet=20
                sheet.cell(row=RowInSheet, column=2, value= str(FGILine3))
                RowInSheet=21
                sheet.cell(row=RowInSheet, column=1, value='FGI r8 = ')
                sheet.cell(row=RowInSheet, column=2, value= str(FGI8Txt))
                RowInSheet=22
                sheet.cell(row=RowInSheet, column=1, value='FGIr9 = ')
                sheet.cell(row=RowInSheet, column=2, value= str(FGI9Txt))
                RowInSheet=23
                sheet.cell(row=RowInSheet, column=1, value='FGIr10 = ')
                sheet.cell(row=RowInSheet, column=2, value= str(FGI10Txt))
                RowInSheet=24
                sheet.cell(row=RowInSheet, column=1, value='UTRAN capabilities : ')
                sheet.cell(row=RowInSheet, column=2, value= str(UTRANcapTxt))
                RowInSheet=25
                sheet.cell(row=RowInSheet, column=1, value='GERAN CS capabilities : ')
                sheet.cell(row=RowInSheet, column=2, value= str(GeranCScapTxt))
                RowInSheet=26
                sheet.cell(row=RowInSheet, column=1, value='GERAN PS capabilities : ')
                sheet.cell(row=RowInSheet, column=2, value= str(GeranPScapTxt))

                for i in range(2,28): # Выделяем заголовки строк 2..27 в колонке А
                    i5 = str(i)
                    cellN = 'A'+i5
                    sheet[cellN].font = Font(name=FBold, size=12, underline='none', color=BLACK, bold=True, italic=False)
                # печатаем позже, т.к. не надо выделять под одну гребенку
                RowInSheet=17
                sheet.cell(row=RowInSheet, column=1, value='Feature Group Indicators: ')
                sheet.cell(row=RowInSheet,column=1).font = Font(name=FBold,size=12, underline='none', color=DARKBLUE, bold=True, italic=False)

                for i in range(2,8): # Вывод данных в колонке B в строках 2..7 жирным моноширинным шрифтом
                    i5 = str(i)
                    cellN = 'B'+i5
                    sheet[cellN].font = Font(name=FMono, size=10, underline='none', color=BLACK, bold= True, italic=False)

                for i in range(18,28): # Вывод данных в колонке B в строках 18..27 жирным моноширинным шрифтом
                    i5 = str(i)
                    cellN = 'B'+i5
                    sheet[cellN].font = Font(name=FMono, size=10, underline='none', color=BLACK, bold= True, italic=False)

                RowInSheet=28
                sheet.cell(row=RowInSheet, column=1, value='CA Combinations : ')
                sheet.cell(row=RowInSheet,column=1).font = Font(name=FBold,size=12, underline='none', color=DARKBLUE, bold=True, italic=False)
                sheet.cell(row=RowInSheet, column=2, value= 'Для TDD расчет ожидаемой скорости выполняется для конфигурации TDD #3 (6DL/3UL таймслотов)')
                Menu = ['Combination Nbr', 'Band', 'DL,bits', 'UL,bits', 'DL#', 'UL#', 'MIMO', 'DL bandwith', 'UL bandwith',
                 'DL Throughput', 'UL Throughput']
                RowInSheet=29
                for i in range (11):
                    sheet.cell(row=RowInSheet, column=i+1, value= Menu[i])
                    sheet.cell(row=RowInSheet,column=i+1).font = Font(name=FBold,size=12, underline='none', color=YELLOW, bold=True, italic=False)
                    sheet.cell(row=RowInSheet,column=i+1).fill = PatternFill(fill_type='solid', start_color=DARKBLUE, end_color=DARKBLUE)
                sheet.column_dimensions['A'].width = 32
                sheet.column_dimensions['B'].width = 31
                for i5 in ['C','D','E','F','G']: # узкие колонки
                    sheet.column_dimensions[i5].width = 10
                for i5 in ['H','I','J','K']: # колонки средней ширины
                    sheet.column_dimensions[i5].width = 18
                CCC = [] # таблица комбинаций для поиска
                CCC1 = (str(CCs[0][1]) + str(CCs[0][4]))
                prevComb = CCs[0][0]
                ccinComb = 0
                CurrComb = CCs[0][0]
                CCCinUsed = False
                CCCGreen = False
                clrCC = C_LightYellow
                # Шаблон строки СУММы
                SumComb = ['','','','','','','','','',0,0]
                #        0      1     2       3       4    5    6     7     8     9      10
                # CCs = [Comb#, Band, DLbits, ULbits, DL#, UL#, MIMO, DLBW, ULBW, DLTpt, ULTpt]
                for i in range(Ncarr):
                    if (CCs[i][0] == prevComb): # продолжение существующей комбинации, накапливаем сумму
                        ccinComb+=1
                        # Добавляем текущие данные в строку Суммы текущей коминации:
                        SumComb[9] = SumComb[9]+CCs[i][9]
                        SumComb[10] = SumComb[10]+CCs[i][10]
                        CCC1=CCC1+(str(CCs[i][1])+str(CCs[i][4])) # Band + количество смежных несущих в бэнде, например, '31', '72'
                        CCC.append(CCC1)
                    else: # начало новой комбинации, старая законилась, надо напечатать ее сумму
                        CCCinUsed = CheckCCCinUsed(CCC1, UsedPat) # old CCC1 of all cc in combination
                        CCCGreen = CheckCCCinUsed(CCC1, UsedPatGreen) # for almost used
                        CurrComb = CCs[i][0]
                        CCC1 = (str(CCs[i][1]) + str(CCs[i][4])) # new CCC1 of first cc
                        # Example: Sum of comb.#3 (4ccs)
                        SumComb[0] = 'SUM of comb.# ' + str(prevComb) + ' (' + str(ccinComb) + 'ccs)'
                        RowInSheet+=1  # добавляем row для вывода суммы закончившейся комбинации
                        if CCCinUsed and BandsFilter:
                            clrCC = C_LightYellow
                        else:
                            clrCC = C_LBlue
                        if CCCGreen and BandsFilter:
                            clrCC = C_LightGreen
                        for j in range(11): # вывод суммы закончившейся комбинации
                            FillCellSum(RowInSheet,j+1,SumComb[j], sheet, clrCC)
                        ccinComb = 1
                        SumComb[9] = CCs[i][9]
                        SumComb[10] = CCs[i][10]
                    prevComb = CCs[i][0]
                    RowInSheet+=1  # добавляем row для выводы строки очередной несущей
                    for j in range(11): # Вывод строки несущей
                        FillCellNorm(RowInSheet, j+1, CCs[i][j], sheet)
                # Конец цикла перебора несущих
                RowInSheet+=1  # добавляем row для вывода суммы последней комбинации
                SumComb[0] = 'SUM of comb.# ' + str(prevComb) + ' (' + str(ccinComb) + 'ccs)'
                CCC1 = CCC1 + (str(CCs[i][1]) + str(CCs[i][4]))  # Band + количество смежных несущих в бэнде, например, '31', '72'
                CCCinUsed = CheckCCCinUsed(CCC1, UsedPat)
                if CCCinUsed and BandsFilter:
                    clrCC = C_LightYellow
                else:
                    clrCC = C_LBlue
                if CCCGreen and BandsFilter:
                    clrCC = C_LightGreen
                for j in range(11): # вывод суммы последней комбинации
                    FillCellSum(RowInSheet, j+1, SumComb[j], sheet, clrCC)
                logging.info('Вкладка Cap.Info заполнена')
                book.save(fnX)
                if os.path.exists(fnX):
                    logging.info('Копия файла успешно сохранена в файл %s', fnX)
                else:
                    logging.info('Файл будет сохранён локально, в каталоге программы %s', 'LatestParsed'+ext)
                    book.save('LatestParsed'+ext)
                    if os.path.exists('LatestParsed'+ext):
                        logging.info('Копия файла успешно сохранена локально в файл %s', 'LatestParsed'+ext)
                    else:
                        logging.error('Копию файла сохранить не удалось')
    else:
        print("CC combinations not found", file=fO)
        logger.warning("CC combinations not found")