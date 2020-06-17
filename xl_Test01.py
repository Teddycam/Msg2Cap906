import openpyxl  # Подключаем библиотеку
# import openpyxl.styles
# from openpyxl import Workbook
# from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill  # Подключаем стили для текста
from openpyxl.styles import colors  # Подключаем цвета для текста и ячеек
from openpyxl.styles import PatternFill  # Подключаем стили для ячеек

#wb = openpyxl.Workbook()  # Создали книгу
wb = openpyxl.load_workbook('sample.xlsx')
print('Sheets: ', wb.sheetnames)
work_sheet = wb.create_sheet(title="Testsheet")  # Создали лист с названием и сделали его активным

     # ws1 = wb.create_sheet("Parsed_Capabilities")                   # insert sheet at the end (by default)
    # sheet = wb["Testsheet"]
    # sheet['A3'] = 'Test'
work_sheet['A1'] = 'Test text'
work_sheet_a1 = work_sheet['A1']  # Создали переменную, в которой содержится ячейка A1 с уже имеющимся текстом
work_sheet['B2'] = 'Test test2'
work_sheet_b2 = work_sheet['B2']  # Создали переменную, в которой содержится ячейка B2 с уже имеющимся текстом

work_sheet_a1.font = Font(size=10, underline='none', color=colors.BLACK, bold=False, italic=True)
work_sheet_b2.font = Font(size=14, underline='double', color=colors.YELLOW, bold=True, italic=False)
work_sheet_b2.fill = PatternFill(fill_type='solid', start_color=colors.DARKBLUE, end_color=colors.DARKBLUE)
print('Sheets: ', wb.sheetnames)
wb.save("Copy3.xlsx")
#=======================================================================================================================
# import openpyxl
# book = openpyxl.load_workbook('sample.xlsx')
#
# print('Sheets: ', book.sheetnames)
#
# # assert "Capabilities" in book.sheetnames, 'Вкладка Capabilities в таблице не найдена'
#
# # try:
# #     sheet = book["Capabilitie"]
# # except:
# #     e = 'Вкладка Capabilities в таблице не найдена'
# #     raise ValueError(e)
#
# if "Capabilities" in book.sheetnames:
#     # Меняем вкладку
#     sheet = book["Capabilities"]
# else:
#     print('Вкладка Capabilities в таблице не найдена')
#     exit(0)
# print('Title = ', sheet.title)
# s = []
# row: object
# for row in sheet.iter_rows():
#     s.append([cell.value for cell in row])
# print(s[-3:])
# print ('s length = ', len(s))
#
# # Add new sheet, fill it and save book copy with new sheet
# if "Parsed_Capabilities" not in book.sheetnames:
#     ws1 = book.create_sheet("Parsed_Capabilities")                   # insert sheet at the end (by default)
#     sheet = book["Parsed_Capabilities"]
#     sheet.sheet_properties.tabColor = "1072BA"
#     sheet['A1'] = 56
#     sheet['A2'] = 43
#     sheet['A3'] = 'Test'
#     print('Вкладка Cap.Info создана и заполнена')
#     book.save("Copy2.xlsx")
#     print('Копия файла сохранена')
# else:
#     print('Вкладка Cap.Info уже существует')
#
# print('Конец работы')
# exit (0)